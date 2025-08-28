# %%
import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
import os
import traceback
import time
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import rainflow_analysis as rfa
import API1183_v2 as api
import processing as pro

overall_start_time = time.time()

categories = ['KS12', 'KS13', 'KS14']
pipe_tally = r"C:\Users\emman\OneDrive - Softnostics\Projects\100001 - 100025\100004 (Acuren - Southbow Screening Software)\Client Documents\KS12-14 Data Collection - (Fixed Headers).xlsx"
pump_stations = r"C:\Users\emman\OneDrive - Softnostics\Projects\100001 - 100025\100004 (Acuren - Southbow Screening Software)\Client Documents\Pump Stations.xlsx"
output_folder = r"C:\Users\emman\OneDrive - Softnostics\Projects\100001 - 100025\100004 (Acuren - Southbow Screening Software)\Client Documents\Results"

df_PT = pd.read_excel(pipe_tally, sheet_name="Sheet1", header=1, engine='calamine')

def create_contour_plot(df, category, feature_id, output_folder):
    # Create a 2x1 figure having contour line plots of: raw and smooth data
    # Establish the levels as the absolute maximum and minimum of both the raw and smooth radius values
    # levels = np.linspace(min([df.o_radius.min(), df.f_radius.min()]), max([df.o_radius.max(), df.f_radius.max()]), 17)
    levels_raw = np.linspace(df.o_radius.min(), df.o_radius.max(), 17)
    levels_smooth = np.linspace(df.f_radius.min(), df.f_radius.max(), 17)

    fig, axs = plt.subplots(2, 1, figsize=(10, 8), sharex=True)
    # Raw data contour plot
    c1 = axs[0].contourf(df.o_axial, df.o_circ, df.o_radius.T, levels=levels_raw, cmap='viridis')
    cb1 = fig.colorbar(c1, ax=axs[0])
    cb1.set_label("Radius (inches)")
    axs[0].set_ylabel(f"{category} Feature {feature_id}: Raw Data")
    axs[0].set_yticks([])

    # Smoothed data contour plot
    c2 = axs[1].contourf(df.f_axial, df.f_circ, df.f_radius.T, levels=levels_smooth, cmap='viridis')
    cb2 = fig.colorbar(c2, ax=axs[1])
    cb2.set_label("Radius (inches)")
    axs[1].set_xlabel("Axial Position (inches)")
    axs[1].set_ylabel(f"{category} Feature {feature_id}: Smoothed Data")
    axs[1].set_yticks([])

    plt.suptitle(f"Comparison for {category} Feature {feature_id}", fontsize=16)
    plt.tight_layout()
    plt.savefig(os.path.join(output_folder, f"{category}_Feature_{feature_id}_Contour.jpg"), dpi=300)
    plt.close(fig)

def create_smoothed_data_sheet(xlsm_file, df_smoothed, df_feature):
    # Save the df_smoothed as a new sheet in the existing .xlsm file
    wb = openpyxl.load_workbook(xlsm_file, read_only=False, keep_vba=True)
    # Remove the sheet if it already exists (optional, for clean overwrite)
    if "Smoothed Data" in wb.sheetnames:
        wb.remove(wb["Smoothed Data"])
    # Create a new sheet named "Smoothed Data"
    ws = wb.create_sheet(title="Smoothed Data")
    for r in dataframe_to_rows(df_smoothed, index=True, header=True):
        ws.append(r)
    # Create a new sheet named "Pipe Tally" and save the df_feature data
    ws_tally = wb.create_sheet(title="Pipe Tally")
    for r in dataframe_to_rows(df_feature, index=False, header=True):
        ws_tally.append(r)
    # Save and close the document (keep the same name)
    wb.save(filename=xlsm_file)
    wb.close()

for category in categories:
    print(f"Processing category: {category}")
    # Filter the DataFrame for the current category
    df_category = df_PT[df_PT['Section'].astype(str).str.contains(category, case=False, na=False)]
    df_dents = df_category[df_category['Feature Type'] == 'Dent']

    # Load the Pump Stations file
    df_stations = pd.read_excel(pump_stations, sheet_name=category, header=0, engine='calamine')

    # Load the pressure history CSV file in folder "Combined Data" based on matching substring
    pressure_history_path = os.path.join(os.getcwd(), "Combined Data", f"{category}_combined.csv")
    # # Read only the header to get column names, this will be used to give dtype for loading the data
    # column_names = pd.read_csv(pressure_history_path, nrows=0).columns.tolist()
    # dtypes = {col: float for col in column_names[1:]}
    # df_pressure_history = pd.read_csv(pressure_history_path, header=0, dtype=dtypes, parse_dates=[column_names[0]])
    df_pressure_history = pd.read_csv(pressure_history_path, header=0, low_memory=False)

    ####################################################################
    # Perform RLA on each feature of interest
    ####################################################################

    # Iterate through df_dents, and based on the value in column 'AP Measure (m)',
    # find the upstream and downstream stations in df_stations based on column 'Continuous Measure (m)'
    # with the station name in column 'Tag Name'. Then, use the upstream and downstream stations to select
    # the pressure history data from df_pressure_history (the columns should match the Tag Names of the stations).

    results = []

    results_folder = f"{category} Results"
    # results_folder = os.path.join(os.getcwd(), results_folder) + '\\'
    results_folder = os.path.join(output_folder, results_folder) + '\\'
    os.mkdir(results_folder)
    count = 0

    import importlib
    importlib.reload(rfa)

    for idx, dent in df_dents.iterrows():
        count += 1
        start_time = time.time()
        try:
            abs_dist = dent['AP Measure (m)']
            
            # Find upstream station
            upstream_candidates = df_stations[df_stations['Continuous Measure (m)'] <= abs_dist]
            
            # Find downstream station
            downstream_candidates = df_stations[df_stations['Continuous Measure (m)'] >= abs_dist]

            # Check if the pressure history contains the required columns
            if upstream_candidates.empty or downstream_candidates.empty:
                print(f"Could not find stations for Feature {dent['Feature ID']} at {abs_dist} m.")
                continue

            upstream_station = upstream_candidates.iloc[-1]
            downstream_station = downstream_candidates.iloc[0]
            upstream_tag = upstream_station['Tag Name']
            downstream_tag = downstream_station['Tag Name']
            
            # Select pressure history data for the upstream and downstream stations, and convert from kPa to psig
            upstream_pressure = pd.to_numeric(df_pressure_history[upstream_tag], errors='coerce').astype(float).to_numpy() * 0.145038
            downstream_pressure = pd.to_numeric(df_pressure_history[downstream_tag], errors='coerce').astype(float).to_numpy() * 0.145038
            time_data = pd.to_datetime(df_pressure_history['Date-Time']).to_numpy()
            
            # Build the dent dict. Convert to Imperial units if necessary.
            dd = rfa.DentData(
                dent_category = category,
                dent_ID = dent['Feature ID'],
                OD = dent['TCPL NPS'], # Already using inch
                WT = dent['TCPL Nominal Wall Thickness [mm]'] * 0.0393701,  # Convert mm to inch
                SMYS = dent['TCPL SMYS [MPa]'] * 145.038,  # Convert MPa to psi
                MAOP = dent['TCPL MOP/MAOP [kPa]'] * 0.145038,  # Convert kPa to psi
                service_years = (time_data[-1] - time_data[0]) / pd.Timedelta(days=365.25),  # Calculate service years from time data
                M = 3,  # Assume exponent of M = 3.0
                min_range = 5,  # Filter at 5psig
                Lx = dent['AP Measure (m)'] * 3.28084,  # Convert meters to feet
                hx = dent['Dent Elevation [m]'] * 3.28084,  # Convert meters to feet
                SG = 0.84, # Assumption from https://www.engineeringtoolbox.com/specific-gravity-liquid-fluids-d_294.html
                L1 = upstream_station['Continuous Measure (m)'] * 3.28084,  # Convert meters to feet
                L2 = downstream_station['Continuous Measure (m)'] * 3.28084,  # Convert meters to feet
                h1 = upstream_station['Elevation'] * 3.28084,  # Convert meters to feet
                h2 = downstream_station['Elevation'] * 3.28084,  # Convert meters to feet
                D1 = dent['TCPL NPS'],
                D2 = dent['TCPL NPS'],
                confidence = dent['Confidence Level Depth [%]'],
                CPS = True if str(dent['CPS That Could Affect HCA']).strip().lower() == 'cps' else False,   # If dent value is "CPS" then return TRUE, otherwise FALSE
                interaction_weld = dent['Weld Fatigue Interaction?'],
                interaction_corrosion = dent['Corrosion Fatigue Interaction?'],
                dent_depth_percent = dent['Peak Depth [%]'],
                ili_pressure = dent['Dent Location Max Pressure (kPa)'] * 0.145038  # Convert kPa to psig
            )
            # Determine the results path
            results_path = os.path.join(results_folder, f"Feature {dent['Feature ID']}") + '\\'
            os.mkdir(results_path)

            # Perform RLA for liquids
            # print(f"{count:04d} / {df_dents.shape[0]:04d} ({round(time.time() - start_time)}s): Processing Feature {dent['Feature ID']}...")
            SSI, CI, MD49_SSI, *_ = rfa.liquid([upstream_pressure, downstream_pressure], time_data, results_path, dd)
            
            result_dict = {
                'Category': category,
                'Feature ID': dd.dent_ID,
                'Upstream Station': upstream_tag,
                'Downstream Station': downstream_tag,
                'SSI': float(SSI).__round__(2),
                'CI': float(CI).__round__(2),
                'MD49_SSI': float(MD49_SSI).__round__(2)
            }
            results.append(result_dict)

            # Save results to notepad after each iteration
            with open(os.path.join(results_folder, f"{category}_RLA_Results.txt"), "a") as f:
                f.write(f"{result_dict}\n")

            # Print or save the results as needed
            print(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Finished rainflow analysis.")

        except Exception as e:
            print(f"{category} Feature {dent['Feature ID']}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Error processing rainflow analysis: {e}")
            continue

        ######################################################################
        # Perform the Data Smoothing using the newly created Excel document
        ######################################################################

        file_path = os.path.join(results_path, f"{category}_Feature_{dd.dent_ID}_Results.xlsm")
        # If category is KS12, use ILI_format = 'KS12', otherwise use ILI_format = 'KS1314'
        if category == 'KS12':
            ILI_format = 'KS12'
        else:
            ILI_format = 'KS1314'

        try:
            # Create a Process class for each file
            df = pro.Process(file_path, ILI_format, dd.OD, dd.WT, dd.SMYS, dd.dent_ID)
            # Smooth the data
            df.smooth_data()
            # Create a new DataFrame containing smoothed data
            df_smoothed = pd.DataFrame(data=df.f_radius, index=df.f_axial, columns=df.f_circ)
            # Create contour plots for the smoothed data
            create_contour_plot(df_smoothed, category, dd.dent_ID, results_path)
            # Create a new sheet in the existing Excel file for the smoothed data
            create_smoothed_data_sheet(file_path, df_smoothed, df)

            print(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully saved smoothed data to file: '{os.path.basename(file_path)}'")

            # Save results to notepad after each iteration. Keep the notepad in the parent output folder (one directory higher)
            with open(os.path.join(os.path.dirname(results_folder), f"{category}_Smoothing_Results.txt"), "a") as f:
                f.write(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully saved smoothed data to file: '{os.path.basename(file_path)}'\n")

        except Exception as e:
            print(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Error saving smoothed data to file: '{os.path.basename(file_path)}. Error: {e}")
            continue

        ######################################################################
        # Perform the MD-4-9 Processing
        ######################################################################
        try:
            api.process_dent_file(file_path)
            print(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully processed MD49: '{os.path.basename(file_path)}'")
            # Save results to notepad after each iteration. Keep the notepad in the parent output folder (one directory higher)
            with open(os.path.join(os.path.dirname(results_folder), f"{category}_MD49_Results.txt"), "a") as f:
                f.write(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully processed MD49: '{os.path.basename(file_path)}'\n")

        except Exception as e:
            print(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Error processing MD49 for file: '{os.path.basename(file_path)}'. Error: {e}")
            print(traceback.format_exc())
            with open(os.path.join(os.path.dirname(results_folder), f"{category}_MD49_Results.txt"), "a") as f:
                f.write(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Error processing MD49 for file: '{os.path.basename(file_path)}'. Error: {e}\n")
                f.write(traceback.format_exc() + "\n")

    # Save the df_results DataFrame to an Excel file
    df_results = pd.DataFrame(results)
    df_results.to_csv(results_path + f"{category}_RLA_Results.csv", index=False)

# %%
