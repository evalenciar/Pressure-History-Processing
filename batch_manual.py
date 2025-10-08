# %%
import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
import matplotlib.image as mpimg
import os
import traceback
import time
import openpyxl
import xlwings as xw
from openpyxl.utils.dataframe import dataframe_to_rows

import rainflow_analysis as rfa
import API1183_v2 as api
import processing as pro

overall_start_time = time.time()

categories = ['KS12', 'KS13', 'KS14']
pipe_tally = r"C:\Users\emman\OneDrive - Softnostics\Projects\100001 - 100025\100004 (Acuren - Southbow Dent Analysis)\Client Documents\KS12-14 Data Collection - (Fixed Headers and ML, Manual).xlsx"
# pump_stations = r"C:\Users\emman\OneDrive - Softnostics\Projects\100001 - 100025\100004 (Acuren - Southbow Dent Analysis)\Client Documents\Pump Stations.xlsx"
caliper_folder = r"C:\Users\emman\OneDrive - Softnostics\Projects\100001 - 100025\100004 (Acuren - Southbow Dent Analysis)\Client Documents\Caliper Radii"
output_folder = r"C:\Users\emman\OneDrive - Softnostics\Projects\100001 - 100025\100004 (Acuren - Southbow Dent Analysis)\Client Documents\Results"
summary_folder = r"C:\Users\emman\OneDrive - Softnostics\Projects\100001 - 100025\100004 (Acuren - Southbow Dent Analysis)\Client Documents\Results\Summary Figures (Manual)"

df_PT = pd.read_excel(pipe_tally, sheet_name="Sheet1", header=1, engine='calamine')

# Add columns to df_PT for Level 0 through Level 3 results
for level in ['Level 0', 'Level 0.5', 'Level 0.5+', 'Level 0.75', 'Level 0.75+', 'Level 1', 'Level 2', 'Level 2 (MD-2-4)', 'Level 3']:
    df_PT[f'{level} w/o SF_eff'] = np.nan
    df_PT[f'{level} w/ SF_eff'] = np.nan
# Add other empty columns for outputs
df_PT['Fatigue Curve'] = np.nan

def find_matching_radius_file(category, target_feature_id):
    # Go through the caliper_folder, check for a subfolder having a matching category, then find a file that contains the feature ID
    for root, dirs, files in os.walk(caliper_folder):
        for dir_name in dirs:
            if category in dir_name:
                for file_path in os.listdir(os.path.join(root, dir_name)):
                    filename = os.path.basename(file_path)
                    feature_id = int(filename.split("Dent")[-1].split(".")[0].split("_")[0].strip(" ").lstrip("0"))
                    if feature_id == target_feature_id:
                        # Get the absolute file path
                        absolute_file_path = os.path.join(root, dir_name, filename)
                        print(f"{category} Feature {target_feature_id}: Found match with: {absolute_file_path}")
                        return absolute_file_path
    raise FileNotFoundError(f"No matching radius file found for {category} Feature {target_feature_id}")

def create_contour_plot(df, category, feature_id, output_folder):
    # Create a 2x1 figure having contour line plots of: raw and smooth data
    # Establish the levels as the absolute maximum and minimum of both the raw and smooth radius values
    # levels = np.linspace(min([df.o_radius.min(), df.f_radius.min()]), max([df.o_radius.max(), df.f_radius.max()]), 17)
    levels_raw = np.linspace(df.o_radius.min(), df.o_radius.max(), 17)
    levels_smooth = np.linspace(df.f_radius.min(), df.f_radius.max(), 17)

    fig, axs = plt.subplots(2, 1, figsize=(10, 8), sharex=True)
    # Raw data contour plot
    c1 = axs[0].contourf(df.o_axial, df.o_circ, df.o_radius.T, levels=levels_raw, cmap='viridis')
    cb1 = fig.colorbar(c1, ax=axs[0])
    cb1.set_label("Radius (inches)")
    axs[0].set_ylabel(f"{category} Feature {feature_id}: Raw Data")
    axs[0].set_yticks([])

    # Smoothed data contour plot
    c2 = axs[1].contourf(df.f_axial, df.f_circ, df.f_radius.T, levels=levels_smooth, cmap='viridis')
    cb2 = fig.colorbar(c2, ax=axs[1])
    cb2.set_label("Radius (inches)")
    axs[1].set_xlabel("Axial Position (inches)")
    axs[1].set_ylabel(f"{category} Feature {feature_id}: Smoothed Data")
    axs[1].set_yticks([])

    plt.suptitle(f"Comparison for {category} Feature {feature_id}", fontsize=16)
    plt.tight_layout()
    plt.savefig(os.path.join(output_folder, f"{category}_Feature_{feature_id}_Contour.png"), dpi=300)
    plt.close(fig)

def create_single_contour_plot(df, category, feature_id, OD, orientation, output_folder, restraint_value: str):
    # Create a single contour line plot of the smoothed data
    # Determine the appropriate restraint_value based on matching substring
    # API 1183 6.5.1.1 Girth Weld Interaction for Fatigue (using values in inches)
    if restraint_value == "Unrestrained":
        a = 0.129
        b = 4.314
        deg = 30
    elif restraint_value == "Deep Restrained" or restraint_value == "Shallow Restrained" or restraint_value == "Restrained":
        a = 0.418
        b = 3.723
        deg = 40

    dc = a * OD + b

    # Add a crosshair at the minimum point, along with text showing the minimum radius value and its axial (x) position
    levels_smooth = np.linspace(df.f_radius.min(), df.f_radius.max(), 17)
    fig, ax = plt.subplots(figsize=(10, 6))
    c = ax.contourf(df.f_axial, df.f_circ, df.f_radius.T, levels=levels_smooth, cmap='viridis')
    cb = fig.colorbar(c, ax=ax)
    min_idx = np.unravel_index(np.argmin(df.f_radius), df.f_radius.shape)
    
    # Add a crosshair at the min_idx. Span the entire plot height and width
    ax.axhline(y=df.f_circ[min_idx[1]], color='red', linestyle='--', linewidth=0.8)
    ax.axvline(x=df.f_axial[min_idx[0]], color='red', linestyle='--', linewidth=0.8)
    # ax.plot(df.f_axial[min_idx[0]], df.f_circ[min_idx[1]], 'ro')  # Red dot at min point
    # Add transparent text box behind the annotation text for better visibility
    # ax.add_patch(plt.Rectangle((df.f_axial[min_idx[0]] - 5, df.f_circ[min_idx[1]] - 5), 80, 40, color='white', alpha=0.5))

    # Place the annotation on the top-left corner of the plot
    # ax.annotate(f"Min Radius: {df.f_radius[min_idx]:.3f} in\nAxial Position: {df.f_axial[min_idx[0]]:.3f} in\nReported Orientation: {orientation}\nCalculated Restraint: {restraint_value}",
    #             xy=(df.f_axial[min_idx[0]], df.f_circ[min_idx[1]]), xytext=(10, ax.get_ylim()[1] - 20),  # Adjust to the top-left corner
    #             textcoords='offset points', fontsize=10, color='black',
    #             bbox=dict(boxstyle="round,pad=0.3", edgecolor='black', facecolor='white', alpha=0.7),
    #             arrowprops=dict(arrowstyle="->", color='black'))
    
    ax.annotate(
        f"Min Radius: {df.f_radius[min_idx]:.3f} in\nAxial Position: {df.f_axial[min_idx[0]]:.3f} in\nReported Orientation: {orientation}\nCalculated Restraint: {restraint_value}",
        xy=(0.02, 0.80),  # Coordinates relative to the axes (2% from the left, 80% from the top)
        xycoords="axes fraction",  # Use axes-relative coordinates
        fontsize=10,
        color="black",
        bbox=dict(boxstyle="round,pad=0.3", edgecolor="black", facecolor="white", alpha=0.7),
    )
    
    # Draw a rectangle centered at the minimum radius point, with width = 2 * dc and height = 2 * deg
    rect = plt.Rectangle((df.f_axial[min_idx[0]] - dc, df.f_circ[min_idx[1]] - deg), 2 * dc, 2 * deg, linewidth=1, edgecolor='black', facecolor='none')
    ax.add_patch(rect)
    cb.set_label("Radius (inches)")
    ax.set_xlabel("Axial Position (inches)")
    ax.set_ylabel("Orientation for Graphical Purposes Only (degrees)")
    # Remove the y-ticks as they do not represent circumferential position
    ax.set_yticks([])
    plt.title(f"{category} Feature {feature_id}", fontsize=16)
    plt.tight_layout()
    plt.savefig(os.path.join(output_folder, f"{category}_Feature_{feature_id}_Smoothed_Contour.png"), dpi=300)
    plt.close(fig)

def create_smoothed_data_sheet(xlsm_file, df_smoothed, df_feature):
    # Save the df_smoothed as a new sheet in the existing .xlsm file
    wb = openpyxl.load_workbook(xlsm_file, read_only=False, keep_vba=True)
    # Remove the sheet if it already exists (optional, for clean overwrite)
    if "Smoothed Data" in wb.sheetnames:
        wb.remove(wb["Smoothed Data"])
    # Create a new sheet named "Smoothed Data"
    ws = wb.create_sheet(title="Smoothed Data")
    for r in dataframe_to_rows(df_smoothed, index=True, header=True):
        ws.append(r)
    # Remove the sheet if it already exists (optional, for clean overwrite)
    if "Pipe Tally" in wb.sheetnames:
        wb.remove(wb["Pipe Tally"])
    # Create a new sheet named "Pipe Tally" and save the df_feature data
    ws_tally = wb.create_sheet(title="Pipe Tally")
    if isinstance(df_feature, pd.Series):
        df_feature = df_feature.to_frame().T
    for r in dataframe_to_rows(df_feature, index=False, header=True):
        ws_tally.append(r)
    # Save and close the document (keep the same name)
    wb.save(filename=xlsm_file)
    wb.close()

def generate_summary_image(overall_results_path, results_path, dd, df, category):
    # Generate a summary image containing the following:
    # 1. General Information table containing: Category, Feature ID, OD, WT, SMYS, Dent Depth (%OD), Dent-Weld Interaction, Dent-Metal Loss Interaction
    # 2. Fatigue Results Summary table containing: select cells from the exported RLA results Excel file
    # 3. Pressure History plot (file normally named "{category}_Feature_{feature_id}_Interpolated_Pressure_History.png")
    # 4. Contour Plot of Smoothed Data (file normally named "{category}_Feature_{feature_id}_Smoothed_Contour.png")
    # 5. Axial Lengths (file normally named "{category}_Feature_{feature_id}_Results_Axial_Lengths.png")
    # 6. Upstream Circumferential Lengths (file normally named "{category}_Feature_{feature_id}_Results_US_Circumferential_Lengths.png")
    # 7. Downstream Circumferential Lengths (file normally named "{category}_Feature_{feature_id}_Results_DS_Circumferential_Lengths.png")

    # Open the RLA excel file and read the required cells for the Fatigue Results table
    rla_file = os.path.join(results_path, f"{dd.dent_category}_Feature_{dd.dent_ID}_Results.xlsm")

    # Use a 'with' statement to ensure the app is closed properly
    with xw.App(visible=False) as app:
        # Open the workbook
        wb = app.books.open(rla_file)

        # Recalculate all formulas in the workbook
        wb.app.calculate()

        # First create the single contour plot using the interaction window criteria
        create_single_contour_plot(df, category, dd.dent_ID, dd.OD, dd.orientation, results_path, wb.sheets['Rainflow'].range('K83').value)

        # Example: 4x2 grid (adjust as needed)
        fig = plt.figure(figsize=(12, 12))
        gs = fig.add_gridspec(4, 2)

        # Main Title
        fig.suptitle(f"{dd.dent_category} Dent {dd.dent_ID} Summary", fontsize=18)

        # 1. General Info Table
        ax0 = fig.add_subplot(gs[0, 0])
        ax0.axis('off')
        data = [
            ["Line", dd.dent_category],
            ["Dent ID", dd.dent_ID],
            ["Pipe OD (in)", dd.OD],
            ["Pipe WT (in)", round(dd.WT, 3)],
            ["Pipe Grade (psi)", 'X70'],
            ["Dent Depth (%OD)", dd.dent_depth_percent],
            ["Dent Orientation [hh:mm]", dd.orientation],
            ["Dent-Weld Interaction", dd.interaction_weld],
            ["Dent-Metal Loss Interaction", dd.interaction_corrosion],
            ["Calculated Restraint", wb.sheets['Rainflow'].range('K83').value],
        ]
        table = ax0.table(cellText=data, loc='center', cellLoc='left')
        # Format the table so first column is dark blue with white text
        for (i, j), cell in table.get_celld().items():
            if j == 0:
                cell.set_facecolor('#003366')  # Dark blue
                cell.set_text_props(color='white', weight='bold')
            else:
                cell.set_facecolor('#f0f0f0')  # Light gray for second column
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        ax0.set_title("General Information", fontsize=14)

        # 2. Fatigue Results Table
        ax1 = fig.add_subplot(gs[1, 0])
        ax1.axis('off')
        fatigue_data = [
            ["Level", 'w/o SF_eff', 'w/ SF_eff'],
            ["Level 0", round(wb.sheets['Summary'].range('C48').value, 1), round(wb.sheets['Summary'].range('D48').value, 1)],
            ["Level 0.5", round(wb.sheets['Summary'].range('C49').value, 1), round(wb.sheets['Summary'].range('D49').value, 1)],
            ["Level 0.5+", round(wb.sheets['Summary'].range('C50').value, 1), round(wb.sheets['Summary'].range('D50').value, 1)],
            ["Level 0.75", round(wb.sheets['Summary'].range('C51').value, 1), round(wb.sheets['Summary'].range('D51').value, 1)],
            ["Level 0.75+", round(wb.sheets['Summary'].range('C52').value, 1), round(wb.sheets['Summary'].range('D52').value, 1)],
            ["Level 1", round(wb.sheets['Summary'].range('C53').value, 1), round(wb.sheets['Summary'].range('D53').value, 1)],
            ["Level 2", round(wb.sheets['Summary'].range('C54').value, 1), round(wb.sheets['Summary'].range('D54').value, 1)],
            ["Level 2 (MD-2-4)", round(wb.sheets['Summary'].range('C55').value, 1), round(wb.sheets['Summary'].range('D55').value, 1)],
            ["Level 3", wb.sheets['Summary'].range('C56').value, wb.sheets['Summary'].range('D56').value]
        ]
        table2 = ax1.table(cellText=fatigue_data, loc='center', cellLoc='left')
        # Format the table so first column and first row is dark blue with white text
        for (i, j), cell in table2.get_celld().items():
            if j == 0 or i == 0:
                cell.set_facecolor('#003366')  # Dark blue
                cell.set_text_props(color='white', weight='bold')
            else:
                cell.set_facecolor('#f0f0f0')  # Light gray for second column
        table2.auto_set_font_size(False)
        table2.set_fontsize(10)
        ax1.set_title("Fatigue Results", fontsize=14)
        wb.close()

    # 3. Pressure History Plot
    ax2 = fig.add_subplot(gs[2, 0])
    img1 = mpimg.imread(os.path.join(results_path,f"{dd.dent_category}_Feature_{dd.dent_ID}_Interpolated_Pressure_History.png"))
    ax2.imshow(img1)
    ax2.axis('off')
    # ax2.set_title("Pressure History", fontsize=14)

    # 4. Smoothed Contour Plot
    ax3 = fig.add_subplot(gs[3, 0])
    img2 = mpimg.imread(os.path.join(results_path,f"{dd.dent_category}_Feature_{dd.dent_ID}_Smoothed_Contour.png"))
    ax3.imshow(img2)
    ax3.axis('off')
    # ax3.set_title("Smoothed Contour", fontsize=14)

    # 5. Axial Lengths Plot
    ax4 = fig.add_subplot(gs[0, 1])
    img3 = mpimg.imread(os.path.join(results_path,f"{dd.dent_category}_Feature_{dd.dent_ID}_Results_Axial_Lengths.png"))
    ax4.imshow(img3)
    ax4.axis('off')
    # ax4.set_title("Axial Lengths", fontsize=14)

    # 6. US Circumferential Lengths Plot
    ax5 = fig.add_subplot(gs[1, 1])
    img4 = mpimg.imread(os.path.join(results_path,f"{dd.dent_category}_Feature_{dd.dent_ID}_Results_US_Circumferential_Lengths.png"))
    ax5.imshow(img4)
    ax5.axis('off')
    # ax5.set_title("Upstream Circumferential Lengths", fontsize=14)

    # 7. DS Circumferential Lengths Plot
    ax6 = fig.add_subplot(gs[2, 1])
    img4 = mpimg.imread(os.path.join(results_path,f"{dd.dent_category}_Feature_{dd.dent_ID}_Results_DS_Circumferential_Lengths.png"))
    ax6.imshow(img4)
    ax6.axis('off')
    # ax6.set_title("Downstream Circumferential Lengths", fontsize=14)

    plt.tight_layout()
    plt.savefig(os.path.join(results_path,f"{dd.dent_category}_Feature_{dd.dent_ID}_Results_Summary.png"), dpi=300)
    plt.savefig(os.path.join(overall_results_path,f"{dd.dent_category}_Feature_{dd.dent_ID}_Results_Summary.png"), dpi=300)
    plt.close(fig)
    
    # Skip the header row of fatigue_data, and convert any negative values to 0. Also convert 'N/A' or None to np.nan
    fatigue_data = fatigue_data[1:]
    for i in range(len(fatigue_data)):
        for j in range(1, len(fatigue_data[i])):
            if isinstance(fatigue_data[i][j], (int, float)) and fatigue_data[i][j] < 0:
                fatigue_data[i][j] = 0
            elif fatigue_data[i][j] == 'N/A' or fatigue_data[i][j] is None:
                fatigue_data[i][j] = np.nan
    return fatigue_data

for category in categories:
    print(f"Processing category: {category}")
    # Filter the DataFrame for the current category
    df_category = df_PT[df_PT['Section'].astype(str).str.contains(category, case=False, na=False)]
    df_dents = df_category[df_category['Feature Type'] == 'Dent']

    # Load the Pump Stations file
    # df_stations = pd.read_excel(pump_stations, sheet_name=category, header=0, engine='calamine')

    # Load the pressure history CSV file in folder "Combined Data" based on matching substring
    # pressure_history_path = os.path.join(os.getcwd(), "Combined Data", f"{category}_combined.csv")
    # # Read only the header to get column names, this will be used to give dtype for loading the data
    # column_names = pd.read_csv(pressure_history_path, nrows=0).columns.tolist()
    # dtypes = {col: float for col in column_names[1:]}
    # df_pressure_history = pd.read_csv(pressure_history_path, header=0, dtype=dtypes, parse_dates=[column_names[0]])
    # df_pressure_history = pd.read_csv(pressure_history_path, header=0, low_memory=False)

    ####################################################################
    # Perform RLA on each feature of interest
    ####################################################################

    # Iterate through df_dents, and based on the value in column 'AP Measure (m)',
    # find the upstream and downstream stations in df_stations based on column 'Continuous Measure (m)'
    # with the station name in column 'Tag Name'. Then, use the upstream and downstream stations to select
    # the pressure history data from df_pressure_history (the columns should match the Tag Names of the stations).

    results = []

    results_folder = f"{category} Results"
    # results_folder = os.path.join(os.getcwd(), results_folder) + '\\'
    results_folder = os.path.join(output_folder, results_folder) + '\\'
    # If the folder already exists, use the existing folder
    if not os.path.isdir(results_folder):
        os.mkdir(results_folder)
    count = 0

    import importlib
    importlib.reload(rfa)
    # %%
    for idx, dent in df_dents.iterrows():
        count += 1
        start_time = time.time()
        try:
            abs_dist = dent['AP Measure (m)']
            
            # Find upstream station
            # upstream_candidates = df_stations[df_stations['Continuous Measure (m)'] <= abs_dist]
            
            # Find downstream station
            # downstream_candidates = df_stations[df_stations['Continuous Measure (m)'] >= abs_dist]

            # Check if the pressure history contains the required columns
            # if upstream_candidates.empty or downstream_candidates.empty:
            #     print(f"Could not find stations for Feature {dent['Feature ID']} at {abs_dist} m.")
            #     continue

            # upstream_station = upstream_candidates.iloc[-1]
            # downstream_station = downstream_candidates.iloc[0]
            # upstream_tag = upstream_station['Tag Name']
            # downstream_tag = downstream_station['Tag Name']
            
            # # Select pressure history data for the upstream and downstream stations, and convert from kPa to psig
            # upstream_pressure = pd.to_numeric(df_pressure_history[upstream_tag], errors='coerce').astype(float).to_numpy() * 0.145038
            # downstream_pressure = pd.to_numeric(df_pressure_history[downstream_tag], errors='coerce').astype(float).to_numpy() * 0.145038
            # time_data = pd.to_datetime(df_pressure_history['Date-Time']).to_numpy()
            
            # Build the dent dict. Convert to Imperial units if necessary.
            dd = rfa.DentData(
                dent_category = category,
                dent_ID = dent['Feature ID'],
                OD = dent['TCPL NPS'], # Already using inch
                WT = dent['TCPL Nominal Wall Thickness [mm]'] * 0.0393701,  # Convert mm to inch
                SMYS = dent['TCPL SMYS [MPa]'] * 145.038,  # Convert MPa to psi
                MAOP = dent['TCPL MOP/MAOP [kPa]'] * 0.145038,  # Convert kPa to psi
                service_years = None,
                M = 3,  # Assume exponent of M = 3.0
                min_range = 5,  # Filter at 5psig
                Lx = dent['AP Measure (m)'] * 3.28084,  # Convert meters to feet
                hx = dent['Dent Elevation [m]'] * 3.28084,  # Convert meters to feet
                SG = 0.84, # Assumption from https://www.engineeringtoolbox.com/specific-gravity-liquid-fluids-d_294.html
                L1 = None,
                L2 = None,
                h1 = None,
                h2 = None,
                D1 = dent['TCPL NPS'],
                D2 = dent['TCPL NPS'],
                confidence = dent['Confidence Level Depth [%]'],
                CPS = True if str(dent['CPS That Could Affect HCA']).strip().lower() == 'cps' else False,   # If dent value is "CPS" then return TRUE, otherwise FALSE
                interaction_weld = dent['Weld Fatigue Interaction?'],
                interaction_corrosion = dent['Corrosion Fatigue Interaction?'],
                dent_depth_percent = dent['Peak Depth [%]'],
                ili_pressure = dent['Dent Location Max Pressure (kPa)'] * 0.145038,  # Convert kPa to psig
                restraint_condition = dent['Fatigue Screening Assumed Restraint Condition'],
                ml_depth_percent = dent['Interacting Peak Depth [%]'],
                ml_location = 'OD' if dent['Interacting Wall Surface'] == 'External' else 'ID' if dent['Interacting Wall Surface'] == 'Internal' else dent['Interacting Wall Surface'],
                orientation = str(dent['Feature Orientation\r\n[hh:mm]'])
            )
            # Determine the results path
            results_path = os.path.join(results_folder, f"Feature {dent['Feature ID']}") + '\\'
            if not os.path.isdir(results_path):
                os.mkdir(results_path)

            # Perform RLA for liquids
            # print(f"{count:04d} / {df_dents.shape[0]:04d} ({round(time.time() - start_time)}s): Processing Feature {dent['Feature ID']}...")
            # SSI, CI, MD49_SSI, *_ = rfa.liquid([upstream_pressure, downstream_pressure], time_data, results_path, dd)

            file_path = os.path.join(results_path, f"{category}_Feature_{dent['Feature ID']}_Results.xlsm")
            wb = openpyxl.load_workbook(filename=file_path, read_only=False, keep_vba=True)

            wbs = wb['Rainflow']
            wbs['K67'] = str(dd.interaction_corrosion)
            wbs['K69'] = str(dd.ml_location)
            wbs['K70'] = str(dd.interaction_weld)
            wbs['K72'] = float(dd.confidence)
            wbs['K73'] = str(dd.CPS)
            wbs['K84'] = str(dd.restraint_condition)

            # Save the resultant Excel workbook into the designated folder
            wb.save(filename=file_path)
            wb.close()
            
            # result_dict = {
            #     'Category': category,
            #     'Feature ID': dd.dent_ID,
            #     'Upstream Station': upstream_tag,
            #     'Downstream Station': downstream_tag,
            #     'SSI': float(SSI).__round__(2),
            #     'CI': float(CI).__round__(2),
            #     'MD49_SSI': float(MD49_SSI).__round__(2)
            # }
            # results.append(result_dict)

            # # Save results to notepad after each iteration
            with open(os.path.join(results_folder, f"{category}_RLA_Results_Manual.txt"), "a") as f:
                # f.write(f"{result_dict}\n")
                f.write(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully saved RLA data to file: '{os.path.basename(file_path)}'\n")

            # Print or save the results as needed
            print(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Finished rainflow analysis.")

        except Exception as e:
            print(f"{category} Feature {dent['Feature ID']}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Error processing rainflow analysis: {e}")
            print(traceback.format_exc())
            continue

        time.sleep(0.1)
        # %%
        ######################################################################
        # Perform the Data Smoothing using the newly created Excel document
        ######################################################################

        file_path = os.path.join(results_path, f"{category}_Feature_{dd.dent_ID}_Results.xlsm")
        # If category is KS12, use ILI_format = 'KS12', otherwise use ILI_format = 'KS1314'
        if category == 'KS12':
            ILI_format = 'KS12'
        else:
            ILI_format = 'KS1314'

        try:
            # Create a Process class for each file
            ili_file_path = find_matching_radius_file(category, dd.dent_ID)
            df = pro.Process(ili_file_path, ILI_format, dd.OD, dd.WT, dd.SMYS, dd.dent_ID)
            # Smooth the data
            df.smooth_data()
            # Create a new DataFrame containing smoothed data
            df_smoothed = pd.DataFrame(data=df.f_radius, index=df.f_axial, columns=df.f_circ)
            # Create contour plots for the smoothed data
            create_contour_plot(df, category, dd.dent_ID, results_path)
            print(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully created contour plots.")
            # Create a new sheet in the existing Excel file for the smoothed data
            create_smoothed_data_sheet(file_path, df_smoothed, dent)

            print(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully saved smoothed data to file: '{os.path.basename(file_path)}'")

            # Save results to notepad after each iteration. Keep the notepad in the parent output folder (one directory higher)
            with open(os.path.join(os.path.dirname(results_folder), f"{category}_Smoothing_Results_Manual.txt"), "a") as f:
                f.write(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully saved smoothed data to file: '{os.path.basename(file_path)}'\n")

        except Exception as e:
            print(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Error saving smoothed data to file: '{os.path.basename(file_path)}. Error: {e}")
            print(traceback.format_exc())
            continue

        time.sleep(0.1)
        # %%
        ######################################################################
        # Perform the MD-4-9 Processing
        ######################################################################
        try:
            api.process_dent_file(file_path, dd.OD, dd.WT)
            print(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully processed MD49: '{os.path.basename(file_path)}'")
            # Save results to notepad after each iteration. Keep the notepad in the parent output folder (one directory higher)
            with open(os.path.join(os.path.dirname(results_folder), f"{category}_MD49_Results_Manual.txt"), "a") as f:
                f.write(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully processed MD49: '{os.path.basename(file_path)}'\n")

        except Exception as e:
            print(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Error processing MD49 for file: '{os.path.basename(file_path)}'. Error: {e}")
            print(traceback.format_exc())
            with open(os.path.join(os.path.dirname(results_folder), f"{category}_MD49_Results_Manual.txt"), "a") as f:
                f.write(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Error processing MD49 for file: '{os.path.basename(file_path)}'. Error: {e}\n")
                f.write(traceback.format_exc() + "\n")

        # Generate the summary image
        try:
            fatigue_data = generate_summary_image(summary_folder, results_path, dd, df, category)
            print(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully generated summary image.")
            # Save results to notepad after each iteration. Keep the notepad in the parent output folder (one directory higher)
            with open(os.path.join(os.path.dirname(results_folder), f"{category}_Summary_Image.txt"), "a") as f:
                f.write(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully generated summary image: '{os.path.basename(file_path)}'\n")
        except Exception as e:
            print(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Error generating summary image: {e}")
            print(traceback.format_exc())
            with open(os.path.join(os.path.dirname(results_folder), f"{category}_Summary_Image.txt"), "a") as f:
                f.write(f"{category} Feature {dd.dent_ID}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Error generating summary image: {e}\n")
                f.write(traceback.format_exc() + "\n")

        # Update df_PT with the results for the current feature
        for level, row in zip(
            ['Level 0', 'Level 0.5', 'Level 0.5+', 'Level 0.75', 'Level 0.75+', 'Level 1', 'Level 2', 'Level 2 (MD-2-4)', 'Level 3'],
            fatigue_data
        ):
            # Locate based on Feature ID, Section = Category, and Feature Type = 'Dent'
            df_PT.loc[(df_PT['Feature ID'] == dd.dent_ID) & (df_PT['Section'].astype(str).str.contains(category, case=False, na=False)) & (df_PT['Feature Type'] == 'Dent'), f'{level} w/o SF_eff'] = row[1]
            df_PT.loc[(df_PT['Feature ID'] == dd.dent_ID) & (df_PT['Section'].astype(str).str.contains(category, case=False, na=False)) & (df_PT['Feature Type'] == 'Dent'), f'{level} w/ SF_eff'] = row[2]

        # Write the results to a .txt file after each iteration
        with open(os.path.join(output_folder, f"All_Fatigue_Results.txt"), "a") as f:
            f.write(f"{category} Feature {dent['Feature ID']}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s)\n")
            for level, row in zip(
                ['Level 0', 'Level 0.5', 'Level 0.5+', 'Level 0.75', 'Level 0.75+', 'Level 1', 'Level 2', 'Level 2 (MD-2-4)', 'Level 3'],
                fatigue_data
            ):
                f.write(f"  {level}: w/o SF_eff = {row[1]}, w/ SF_eff = {row[2]}\n")
            f.write("\n")

        # Attempt to save the updated df_PT to an Excel file after each iteration
        live_excel_path = os.path.join(output_folder, "Updated_Pipe_Tally_Live (Please Close Me).xlsx")
        try:
            df_PT.to_excel(live_excel_path, index=False)
            print(f"Live update saved to {live_excel_path}")
        except Exception as e:
            print(f"Could not save live update to {live_excel_path}: {e}")

    # Save the df_results DataFrame to an Excel file
    df_results = pd.DataFrame(results)
    df_results.to_csv(results_path + f"{category}_RLA_Results.csv", index=False)

# Save the updated df_PT to an Excel file
output_excel_path = os.path.join(output_folder, "Updated_Pipe_Tally.xlsx")
df_PT.to_excel(output_excel_path, index=False)
print(f"Updated df_PT saved to {output_excel_path}")
# %%
