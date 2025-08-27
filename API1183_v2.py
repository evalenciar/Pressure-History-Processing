# -*- coding: utf-8 -*-
"""
Created on Fri Aug 15 13:42:42 2025

@author: ahassanin, evalencia

Version 0.2.0 - adds batch processing and slope change selection

This script references the API1183.py Version 0.1.18 script provided by ahassanin with some 
minor modifications to work with batch processing and slope change selection.
"""

from pathlib import Path
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.legend import Legend
from openpyxl.styles.colors import Color
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font
import math
import numpy as np
from scipy.signal import savgol_filter
from openpyxl.utils.dataframe import dataframe_to_rows

palette = [
    '#1f77b4',  # muted blue
    '#ff7f0e',  # safety orange
    '#2ca02c',  # cooked asparagus green
    '#d62728',  # brick red
    '#9467bd',  # muted purple
    '#8c564b',  # chestnut brown
    '#e377c2',  # raspberry yogurt pink
    '#7f7f7f',  # middle gray
    '#bcbd22',  # curry yellow‑green
    '#17becf',  # blue‑teal
    '#aec7e8',  # light pastel blue
    '#ffbb78'   # light pastel orange
]

def _find_deflection_initiation(data, nominal_radius, window_size=11, slope_tolerance=0.001, closeness_percentage=5, buffer_offset=5, min_consecutive_deviations=3, circumferential_mode=False):
    """
    Internal helper to find the initiation index of deflection assuming the data starts from pristine pipe and moves towards the dent.
    
    In default mode, scans from the left to find the first point where the condition deviates from pristine (flat slope and radius close to nominal) 
    in a sustained manner (for at least min_consecutive_deviations points) to handle noise and vibrations in the pristine section.
    
    In circumferential_mode, identifies the first local maximum (slope ≈ 0) where the radius is close to nominal_radius, suitable for cases where a peak near nominal marks the deflection point (e.g., due to interacting features).
    
    Parameters
    ----------
    data : np.ndarray
        Array of caliper radius measurements along the pipeline section. Must be a 1D array of floating-point values representing radii at sequential points.
        - Guidance: Ensure the data is preprocessed to represent either the inbound (pristine to dent) or reversed outbound (dent to pristine) half. Array length should be at least `window_size`. Smooth extreme outliers manually if necessary, as the function handles moderate noise via the Savitzky-Golay filter.
        - Example: `np.array([10.0, 9.99, 9.98, ..., 8.5])` where 10.0 is nominal radius.
    nominal_radius : float
        Expected radius of the pristine (undamaged) pipe, used as the baseline for detecting deviations in radius and slope.
        - Guidance: Use the pipeline's nominal inner radius from design specs or compute as the mean of a known pristine section (e.g., first/last 10–20 points if flat). Adjust empirically if data shows a sensor bias by inspecting the data plot.
        - Example: 150.0 mm for a medium-sized pipeline.
    window_size : int, optional (default=11)
        Window length for the Savitzky-Golay filter to compute a smoothed first derivative (slope), reducing noise effects.
        - Guidance: Must be an odd integer > 1 (e.g., 5, 7, 11, 15). Smaller windows (5–7) preserve local details but are noise-sensitive; larger windows (15–21) smooth more, suitable for noisy data but may delay detection of sharp changes. Choose based on data resolution: larger for high sampling rates, smaller for sparse data. Test by plotting the smoothed derivative.
        - Example: 11 for moderately noisy data with 100+ points.
    slope_tolerance : float, optional (default=0.001)
        Threshold for considering the slope (first derivative) as "flat" (close to zero), indicating a pristine section or a local maximum in circumferential_mode.
        - Guidance: Units are radius change per data point (e.g., mm/sample). Set based on expected noise: 0.001–0.01 for low noise, 0.01–0.1 for high noise. In circumferential_mode, this defines the zero-slope threshold for local maxima. Inspect the derivative of pristine sections to estimate typical slope variations.
        - Example: 0.001 for stable data, 0.05 for noisy data.
    closeness_percentage : float, optional (default=5)
        Percentage tolerance for how close the radius must be to `nominal_radius` to be considered pristine or a valid local maximum.
        - Guidance: Set based on expected radius variation in pristine sections due to noise or manufacturing tolerances. 2–5% is typical for well-calibrated sensors; increase to 10% for noisy data or variable pipe conditions. Check radius histograms of pristine sections to confirm.
        - Example: 5 for ±5% tolerance around nominal radius.
    buffer_offset : int, optional (default=5)
        Number of points to subtract from the detected initiation index to include preceding pristine data, ensuring no relevant data is excluded.
        - Guidance: Use 3–10 points depending on data resolution and desired margin. Smaller offsets (3–5) for high-resolution data; larger (5–10) for sparse or noisy data to capture context before deflection. Ensure it doesn’t push the index below 0.
        - Example: 5 points for moderate resolution.
    min_consecutive_deviations : int, optional (default=3)
        In default mode, minimum number of consecutive points that must deviate from pristine conditions to confirm the start of deflection. In circumferential_mode, minimum consecutive points with near-zero slope and radius close to nominal to confirm a local maximum.
        - Guidance: Use 2–5 points; 2 for sharp deflections or clear maxima, 3–5 for noisy data to avoid false positives from brief spikes. Adjust based on inspection of noise patterns or slope behavior near maxima.
        - Example: 3 for typical pipeline data with moderate noise.
    circumferential_mode : bool, optional (default=False)
        If True, detects the deflection point as the global maximum (highest radius) among points with slope ≈ 0 (within slope_tolerance) and radius close to nominal_radius (within closeness_percentage). If False, uses the original sustained-deviation approach.
        - Guidance: Enable for cases where a circumferential feature or interacting dent causes a prominent peak near nominal radius at the deflection boundary, and the global maximum is the most significant marker. Disable for standard dent detection where sustained radius/slope deviation marks the start. Inspect data plots to confirm a prominent peak near nominal radius.
        - Example: True for circumferential feature analysis, False for typical dents.
    
    Returns
    -------
    init_idx : int or None
        The index where deflection initiates (with buffer applied), or None if no deflection detected.
    """
    if len(data) < window_size:
        raise ValueError("Data length must be at least the window size.")
    
    # Compute smoothed first derivative
    der = savgol_filter(data, window_length=window_size, polyorder=2, deriv=1)
    
    # Closeness tolerance for radius
    closeness_tol = (closeness_percentage / 100) * nominal_radius
    
    # Define pristine condition: flat slope and radius close to nominal
    def is_pristine(i):
        return abs(der[i]) <= slope_tolerance and abs(data[i] - nominal_radius) <= closeness_tol
    
    start_idx = 0
    if circumferential_mode:
        # Find global maximum among points with near-zero slope and radius close to nominal
        max_radius = -np.inf
        max_idx = None
        i = 0
        while i < len(der):
            if abs(der[i]) <= slope_tolerance and abs(data[i] - nominal_radius) <= closeness_tol:
                # Check for consecutive points to confirm the maximum
                consec = 1
                j = i + 1
                while j < len(der) and abs(der[j]) <= slope_tolerance and abs(data[j] - nominal_radius) <= closeness_tol:
                    consec += 1
                    j += 1
                if consec >= min_consecutive_deviations:
                    # Check if this region contains a higher radius
                    region_max = np.max(data[i:j])
                    if region_max > max_radius:
                        max_radius = region_max
                        max_idx = i + np.argmax(data[i:j])
                i = j
            else:
                i += 1
        start_idx = max_idx
    else:
        # Default mode: find sustained deviation from pristine
        i = 0
        while i < len(der):
            if is_pristine(i):
                start_idx = i + 1
                i += 1
            else:
                # Check for consecutive deviations
                consec = 1
                j = i + 1
                while j < len(der) and not is_pristine(j):
                    consec += 1
                    j += 1
                if consec >= min_consecutive_deviations:
                    break
                else:
                    # Skip the noise blip and continue
                    start_idx = j
                    i = j
    
    # If no valid deflection point found (entire data pristine or no valid maximum)
    if start_idx >= len(data):
        return None
    
    # Apply buffer offset (subtract for initiation to include more preceding data)
    start_idx = max(0, start_idx - buffer_offset)
    
    return start_idx

def find_inbound_deflection_start(inbound_data, nominal_radius, window_size=11, slope_tolerance=0.001, closeness_percentage=5, buffer_offset=5, min_consecutive_deviations=3, circumferential_mode=False):
    """
    Finds the index where the caliper begins deflecting in the inbound half of the data.
    
    The inbound_data should be the half starting from pristine pipe (with noise and vibrations) moving towards the deepest part of the dent (minimum radius).
    
    Uses a slope approach to detect either sustained deviation from pristine conditions (default) or a local maximum near nominal radius (circumferential_mode).
    
    Parameters
    ----------
    inbound_data : np.ndarray
        Array of caliper radius measurements for the inbound half, from pristine to dent minimum.
        - Guidance: Split full data at the minimum radius index (e.g., `inbound_data = full_data[:min_idx + 1]`). Length must be >= `window_size`. Smooth outliers if extreme.
        - Example: `np.array([10.0, 9.99, 9.98, ..., 8.5])`.
    nominal_radius : float
        Expected radius of the pristine pipe.
        - Guidance: Use design specs or mean of pristine section (e.g., first 10–20 points). Adjust for sensor bias if needed.
        - Example: 150.0 mm.
    window_size : int, optional (default=11)
        Window for Savitzky-Golay filter to smooth derivatives.
        - Guidance: Odd integer (5–21). Use 5–7 for sharp changes, 15–21 for noisy data. Match to data resolution.
        - Example: 11.
    slope_tolerance : float, optional (default=0.001)
        Slope threshold for pristine sections or local maxima.
        - Guidance: 0.001–0.01 for low noise, 0.01–0.1 for high noise. In circumferential_mode, defines zero-slope for maxima.
        - Example: 0.001.
    closeness_percentage : float, optional (default=5)
        Percentage tolerance for radius closeness to nominal.
        - Guidance: 2–5% for stable data, 5–10% for noisy or variable pipes. Verify with radius histograms.
        - Example: 5.
    buffer_offset : int, optional (default=5)
        Points to subtract from start index for preceding data.
        - Guidance: 3–10 points; smaller for high-resolution, larger for sparse/noisy data.
        - Example: 5.
    min_consecutive_deviations : int, optional (default=3)
        Minimum consecutive deviations (default mode) or consecutive points at local maximum (circumferential_mode).
        - Guidance: 2–5; 2 for sharp deflections/maxima, 3–5 for noisy data.
        - Example: 3.
    circumferential_mode : bool, optional (default=False)
        If True, detects deflection at the first local maximum (slope ≈ 0) near nominal radius. If False, uses sustained deviation.
        - Guidance: Enable for circumferential features causing peaks near nominal radius. Inspect data for such maxima.
        - Example: True for circumferential analysis.
    
    Returns
    -------
    start_idx : int or None
        The relative index in inbound_data where deflection starts (with buffer), or None if not detected.
    """
    if circumferential_mode:
        # Reverse the data to treat it as starting from pristine to dent
        inbound_data = inbound_data[::-1]
    init_idx = _find_deflection_initiation(inbound_data, nominal_radius, window_size, slope_tolerance, closeness_percentage, buffer_offset, min_consecutive_deviations, circumferential_mode)

    if circumferential_mode:
        # Convert back to original outbound index (this is the end index)
        init_idx = len(inbound_data) - 1 - init_idx

        # Cap to data length
        init_idx = min(len(inbound_data) - 1, init_idx)

    return init_idx

def find_outbound_deflection_end(outbound_data, nominal_radius, window_size=11, slope_tolerance=0.001, closeness_percentage=5, buffer_offset=5, min_consecutive_deviations=3, circumferential_mode=False):
    """
    Finds the index where the caliper ends deflecting in the outbound half of the data.
    
    The outbound_data should be the half starting from the deepest part of the dent (minimum radius) and moving back towards pristine pipe.
    
    Internally reverses the data to apply the same slope-based initiation detection, adjusting for direction.
    
    Parameters
    ----------
    outbound_data : np.ndarray
        Array of caliper radius measurements for the outbound half, from dent minimum to pristine.
        - Guidance: Split full data at minimum radius (e.g., `outbound_data = full_data[min_idx:]`). Length must be >= `window_size`. Smooth outliers if extreme.
        - Example: `np.array([8.5, ..., 9.95, 9.98, 10.0])`.
    nominal_radius : float
        Expected radius of the pristine pipe.
        - Guidance: Same as for inbound; use consistent value across both functions.
        - Example: 150.0 mm.
    window_size : int, optional (default=11)
        Window for Savitzky-Golay filter.
        - Guidance: Same as for inbound; keep consistent for symmetry.
        - Example: 11.
    slope_tolerance : float, optional (default=0.001)
        Slope threshold for pristine sections or local maxima.
        - Guidance: Same as for inbound; use same value for consistency.
        - Example: 0.001.
    closeness_percentage : float, optional (default=5)
        Percentage tolerance for radius closeness.
        - Guidance: Same as for inbound; use same value.
        - Example: 5.
    buffer_offset : int, optional (default=5)
        Points effectively added to end index for following data (due to reversal).
        - Guidance: Same as for inbound; applies as a buffer in the reversed data context.
        - Example: 5.
    min_consecutive_deviations : int, optional (default=3)
        Minimum consecutive deviations or points at local maximum in reversed data.
        - Guidance: Same as for inbound; use same value.
        - Example: 3.
    circumferential_mode : bool, optional (default=False)
        If True, detects deflection end at the first local maximum near nominal radius in reversed data. If False, uses sustained deviation.
        - Guidance: Enable for circumferential features. Use same setting as inbound for consistency.
        - Example: True for circumferential analysis.
    
    Returns
    -------
    end_idx : int or None
        The relative index in outbound_data where deflection ends (with buffer), or None if not detected.
    """
    if len(outbound_data) < window_size:
        raise ValueError("Data length must be at least the window size.")
    
    # Reverse the data to treat it as starting from pristine to dent
    reversed_data = outbound_data[::-1]
    
    # Find initiation in reversed data
    rev_init_idx = _find_deflection_initiation(reversed_data, nominal_radius, window_size, slope_tolerance, closeness_percentage, buffer_offset, min_consecutive_deviations, circumferential_mode)
    
    if rev_init_idx is None:
        return None
    
    # Convert back to original outbound index (this is the end index)
    end_idx = len(outbound_data) - 1 - rev_init_idx
    
    # Cap to data length
    end_idx = min(len(outbound_data) - 1, end_idx)
    
    return end_idx

def determine_nominal(df: pd.Series, expected_nominal: float, threshold: float) -> float:
    # Use the first and last foot of axial data to determine an average nominal radius
    # The index is the axial displacement in inches
    first_set = df[df.index <= 12]
    last_set = df[df.index >= (df.index.max() - 12)]
    avg_nominal = ((first_set.mean() + last_set.mean()) / 2).mean()

    # Then compare this calculated average nominal radius to the expected nominal radius
    if abs(avg_nominal - expected_nominal) <= (expected_nominal * threshold / 100):
        # If the difference is within a threshold percentage, then use the expected nominal radius
        return expected_nominal
    # If the difference is within a threshold percentage, then use the expected nominal radius, otherwise use the calculated average
    return avg_nominal

def process_dent_file(file_path: Path, sheet_name: str="Smoothed Data"):
    # Defaults
    baseline_axial = 0.025
    baseline_circum = -0.15

    # Load existing Excel document and read data from the "Smoothed Data" sheet, and orient so Theta in columns, Z in rows
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=[1], index_col=0)

    # Find deepest dent point
    z_min, theta_min = df.stack().idxmin()
    # theta_min, z_min = df.stack().idxmin()
    axial_profile   = df[theta_min]
    circum_profile  = df.loc[z_min]

    # Split axial data into US/DS
    ax_pos      = np.argmin(axial_profile.values)
    axial_up    = axial_profile.iloc[:ax_pos+1]
    axial_down  = axial_profile.iloc[ax_pos:]
    max_dist = math.ceil(axial_profile.index.max()/50)*50
    
    # Split circum data to CCW/CW
    ce_pos      = np.argmin(circum_profile.values)
    circum_ccw  = circum_profile.iloc[:ce_pos+1]
    circum_cw = circum_profile.iloc[ce_pos:]

    # DataFrames
    axial_df   = pd.DataFrame({'Upstream': axial_up, 'Downstream': axial_down})
    circum_df  = pd.DataFrame({'Counterlockwise': circum_ccw, 'Clockwise': circum_cw})

    # Open the file_path Excel document and add new sheets
    wb = openpyxl.load_workbook(file_path, read_only=False, keep_vba=True)

    # Load the metadata from the Summary tab
    metadata = {
        'Pipe OD' : wb['Summary']['D2'].value,  # Outer diameter (inches)
        'WT' : wb['Summary']['D3'].value,       # Wall thickness (inches)
        'SMYS' : wb['Summary']['D4'].value,     # Specified minimum yield strength (psi)
    }

    # Create a new sheet titled "Axial Profiles", and assing to sht
    axial_name = 'Axial Profiles'
    sht = wb.create_sheet(title=axial_name)

    # Write the axial_df (including index and headers) to sht without using writer to preserve Excel VBA macros
    for r_idx, row in enumerate(dataframe_to_rows(axial_df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            sht.cell(row=r_idx, column=c_idx, value=value)

    # --- 3.1 Compute axial baseline by walking & interpolating ---
    
    # Upstream
    nominal_default    = (metadata['Pipe OD'] / 2) - metadata['WT']
    nominal_IR = determine_nominal(df, nominal_default, 5)
    nominal_us = nominal_IR
    rads_us       = axial_up.values
    dists_us      = axial_up.index.values
    dent_depth_us = nominal_us - rads_us.min()
    baseline_us_default   = nominal_us - baseline_axial * dent_depth_us
    rads_us_idx = find_inbound_deflection_start(rads_us, nominal_us, slope_tolerance=0.004)
    baseline_us = rads_us[rads_us_idx] if rads_us_idx is not None else baseline_us_default
    # Re-establish the dent_depth_us value based on the new baseline_us value
    dent_depth_us = baseline_us - rads_us.min()

    # find peak index in upstream
    peak_idx_us = np.argmin(rads_us)
    i = peak_idx_us
    # walk backward until rads_us[i] >= baseline_us
    while i > 0 and rads_us[i] < baseline_us:
        i -= 1
    
    if i == peak_idx_us:
        # never recovered → use peak location
        baseline_dist_us = dists_us[peak_idx_us]
    else:
        # interpolate between points i and i+1
        r0, x0 = rads_us[i],   dists_us[i]
        r1, x1 = rads_us[i+1], dists_us[i+1]
        if not np.isfinite(r1 - r0) or r1 == r0:
            baseline_dist_us = 0.5 * (x0 + x1)
        else:
            frac = (baseline_us - r0) / (r1 - r0)
            baseline_dist_us = x0 + frac * (x1 - x0)
    
    peak_dist_us = dists_us[peak_idx_us]
        
    # Downstream
    nominal_ds = nominal_IR
    rads_ds       = axial_down.values
    dists_ds      = axial_down.index.values
    dent_depth_ds = nominal_ds - rads_ds.min()
    baseline_ds_default   = nominal_ds - baseline_axial * dent_depth_ds
    rads_ds_idx = find_outbound_deflection_end(rads_ds, nominal_ds, slope_tolerance=0.004)
    baseline_ds = rads_ds[rads_ds_idx] if rads_ds_idx is not None else baseline_ds_default
    # Re-establish the dent_depth_ds value based on the new baseline_ds value
    dent_depth_ds = baseline_ds - rads_ds.min()

    # find peak index in downstream (same global peak)
    peak_idx_ds = np.argmin(rads_ds)
    i = peak_idx_ds
    # walk forward until rads_ds[i] >= baseline_ds
    while i < len(rads_ds) - 1 and rads_ds[i] < baseline_ds:
        i += 1
    
    if i == peak_idx_ds:
        baseline_dist_ds = dists_ds[peak_idx_ds]
    else:
        # interpolate between points i-1 and i
        r0, x0 = rads_ds[i-1], dists_ds[i-1]
        r1, x1 = rads_ds[i],   dists_ds[i]
        if not np.isfinite(r1 - r0) or r1 == r0:
            baseline_dist_ds = 0.5 * (x0 + x1)
        else:
            frac = (baseline_ds - r0) / (r1 - r0)
            baseline_dist_ds = x0 + frac * (x1 - x0)
        
    peak_dist_ds = dists_ds[peak_idx_ds]
        
    # LAX table header - upstream
    sht['E27'] = 'Upstream Lengths'
    sht['E28'] = 'Max Dent Depth'; sht['G28'] = dent_depth_us
    sht['E29'] = 'Baseline';       sht['F29'] = baseline_dist_us; sht['G29'] = baseline_us

    # LAX table header - downstream
    sht['K27'] = 'Downstream Lengths'
    sht['K28'] = 'Max Dent Depth'; sht['M28'] = dent_depth_ds
    sht['K29'] = 'Baseline';       sht['L29'] = baseline_dist_ds; sht['M29'] = baseline_ds
        
    # Prepare LAX% & endpoints - upstream
    percentages_us = [95, 90, 85, 75, 60, 50, 40, 30, 20, 15, 10, 5]
    distances_us   = axial_up.index.values

    horiz_start = 45
    base_row    = horiz_start - 1  # for baseline endpoints
        
    # Prepare LAX% & endpoints - downstream
    percentages_ds = [95, 90, 85, 75, 60, 50, 40, 30, 20, 15, 10, 5]
    distances_ds   = axial_down.index.values

    horiz_start = 45
    base_row    = horiz_start - 1  # for baseline endpoints

    # baseline points - upstream
    sht[f'E{base_row}'] = peak_dist_us
    sht[f'F{base_row}'] = baseline_dist_us
    sht[f'G{base_row}'] = baseline_us
    sht[f'H{base_row}'] = baseline_us

    targets_us = []
    for i, p in enumerate(percentages_us):
        # target_r = baseline_us - (p/100) * dent_depth_us
        target_r = rads_us.min() + (1 - p/100) * dent_depth_us
    
        # walk *backward* from the dent peak until radius >= target
        j = peak_idx_us
        while j > 0 and rads_us[j] < target_r:
            j -= 1
    
        if rads_us[j] < target_r:
            # never gets back to the target → use the peak location
            x_at = peak_dist_us
        else:
            # interpolate between j and j+1
            # if j==len−1 just take distances_us[j]
            if j == len(rads_us)-1 or not np.isfinite(rads_us[j+1] - rads_us[j]) or rads_us[j+1] == rads_us[j]:
                x_at = distances_us[j]
            else:
                r0, x0 = rads_us[j],   distances_us[j]
                r1, x1 = rads_us[j+1], distances_us[j+1]
                frac   = (target_r - r0) / (r1 - r0)
                x_at   = x0 + frac * (x1 - x0)
    
        LAX = abs(x_at - peak_dist_us)
        row = 30 + i
        sht[f'E{row}'] = f'LAX{p}%'
        sht[f'F{row}'] = LAX
        sht[f'G{row}'] = target_r

        # chart endpoints
        r = horiz_start + i
        sht[f'E{r}'] = peak_dist_us
        sht[f'F{r}'] = x_at
        sht[f'G{r}'] = target_r
        sht[f'H{r}'] = target_r

        targets_us.append((p, x_at))
            
    # baseline points – downstream  ← add these back in!
    sht[f'K{base_row}'] = peak_dist_ds
    sht[f'L{base_row}'] = baseline_dist_ds
    sht[f'M{base_row}'] = baseline_ds
    sht[f'N{base_row}'] = baseline_ds

    targets_ds = []
    for i, p in enumerate(percentages_ds):
        # target_r = baseline_ds - (p/100) * dent_depth_ds
        target_r = rads_ds.min() + (1 - p/100) * dent_depth_ds
    
        # walk *forward* from the dent peak until radius >= target
        j = peak_idx_ds
        while j < len(rads_ds)-1 and rads_ds[j] < target_r:
            j += 1
    
        if rads_ds[j] < target_r:
            # never reaches target → use peak location
            x_at = peak_dist_ds
        else:
            # interpolate between j-1 and j
            if j == 0 or not np.isfinite(rads_ds[j] - rads_ds[j-1]) or rads_ds[j] == rads_ds[j-1]:
                x_at = distances_ds[j]
            else:
                r0, x0 = rads_ds[j-1], distances_ds[j-1]
                r1, x1 = rads_ds[j],   distances_ds[j]
                frac   = (target_r - r0) / (r1 - r0)
                x_at   = x0 + frac * (x1 - x0)
    
        LAX = abs(x_at - peak_dist_ds)
        row = 30 + i
        sht[f'K{row}'] = f'LAX{p}%'
        sht[f'L{row}'] = LAX
        sht[f'M{row}'] = target_r

        # chart endpoints
        r = horiz_start + i
        sht[f'K{r}'] = peak_dist_ds
        sht[f'L{r}'] = x_at
        sht[f'M{r}'] = target_r
        sht[f'N{r}'] = target_r

        targets_ds.append((p, x_at))

    # Upstream Areas
    sht['E68'] = 'Upstream Areas'
    sht['E69'] = 'Distance'
    sht['F69'] = 'Radius'
    start_us = min(baseline_dist_us, peak_dist_us)
    end_us   = max(baseline_dist_us, peak_dist_us)
    mask_seg_us = (axial_up.index >= start_us) & (axial_up.index <= end_us)
    seg_dist_us = axial_up.index[mask_seg_us].tolist()
    seg_rad_us  = axial_up.values[mask_seg_us].tolist()
    for i, value in enumerate(seg_dist_us):
        sht[f'E{70 + i}'] = value
    for i, value in enumerate(seg_rad_us):
        sht[f'F{70 + i}'] = value

    # Downstream Areas
    sht['K68'] = 'Downstream Areas'
    sht['K69'] = 'Distance'
    sht['L69'] = 'Radius'
    start_ds = min(baseline_dist_ds, peak_dist_ds)
    end_ds   = max(baseline_dist_ds, peak_dist_ds)
    mask_seg_ds = (axial_down.index >= start_ds) & (axial_down.index <= end_ds)
    seg_dist_ds = axial_down.index[mask_seg_ds].tolist()
    seg_rad_ds  = axial_down.values[mask_seg_ds].tolist()
    for i, value in enumerate(seg_dist_ds):
        sht[f'K{70 + i}'] = value
    for i, value in enumerate(seg_rad_ds):
        sht[f'L{70 + i}'] = value

    # Area calc & store in list - upstream
    sht['G69'] = 'Area'
    area_list_us = [0]
    for j in range(1, len(seg_dist_us)):
        d_i, d_im1 = seg_dist_us[j], seg_dist_us[j-1]
        r_i, r_im1 = seg_rad_us[j], seg_rad_us[j-1]
        area_us = 0.5 * (d_i - d_im1) * abs((r_i - baseline_us) + (r_im1 - baseline_us))
        area_list_us.append(area_us)
        sht[f'G{71 + j}'] = area_us

    # Area calc & store in list - downstream
    sht['M69'] = 'Area'
    area_list_ds = [0]
    for j in range(1, len(seg_dist_ds)):
        d_i, d_im1 = seg_dist_ds[j], seg_dist_ds[j-1]
        r_i, r_im1 = seg_rad_ds[j], seg_rad_ds[j-1]
        area_ds = 0.5 * (d_i - d_im1) * abs((r_i - baseline_ds) + (r_im1 - baseline_ds))
        area_list_ds.append(area_ds)
        sht[f'M{71 + j}'] = area_ds

    # AX% cumulative areas - upstream
    cum_row = 30
    for p in [85, 75, 60, 50, 40, 30, 20, 15, 10]:
        x_at_p = next(x for pct, x in targets_us if pct == p)
        idx_p  = min(range(len(seg_dist_us)), key=lambda k: abs(seg_dist_us[k] - x_at_p))
        cum_area = sum(area_list_us[idx_p:])
        sht[f'H{cum_row}'] = f'AAX{p}%'
        sht[f'I{cum_row}'] = cum_area
        cum_row += 1

    # AX% cumulative areas - downstream
    cum_row = 30
    for p in [85,75,60,50,40,30,20,15,10]:
        x_at_p = next(x for pct, x in targets_ds if pct == p)
        idx_p  = min(range(len(seg_dist_ds)), key=lambda k: abs(seg_dist_ds[k] - x_at_p))
        # sum from baseline inward to the p% point
        cum_area = sum(area_list_ds[:idx_p+1])
        sht[f'N{cum_row}'] = f'AAX{p}%'
        sht[f'O{cum_row}'] = cum_area
        cum_row += 1

    # <-------------------------------------------------- REPEAT THIS PART TWICE FOR US and DS CCW and CW COMBINATIONS
    # Circumferential data
    # Create a new sheet titled "Circumferential Profiles", and assing to sht
    circumUS_name = 'CircumUS Profiles'
    sht = wb.create_sheet(title=circumUS_name)

    # Write the circum_df to sht without using writer to preserve Excel VBA macros
    for r_idx, row in enumerate(dataframe_to_rows(circum_df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            sht.cell(row=r_idx, column=c_idx, value=value)
    
    # --- 6.1 Compute interpolated baseline for CCW and CW ---
    
    # Common nominal
    nominal = nominal_IR
    
    # 1) CCW (counterclockwise) – search backward from the dent peak
    rads_ccw, angs_ccw = circum_ccw.values, circum_ccw.index.values
    dent_depth_ccw   = nominal - rads_ccw.min()
    baseline_ccw_default     = nominal - baseline_circum * dent_depth_ccw
    rads_ccw_idx = find_inbound_deflection_start(rads_ccw, nominal, slope_tolerance=0.004, circumferential_mode=True)
    # baseline_ccw = rads_ccw[rads_ccw_idx] if rads_ccw_idx is not None else baseline_ccw_default
    baseline_ccw = baseline_us
    # Re-establish the dent_depth_ccw value based on the new baseline_ccw value
    dent_depth_ccw = baseline_ccw - rads_ccw.min()

    # find peak index
    peak_idx_ccw     = np.argmin(rads_ccw)
    # walk backward until we reach or exceed the baseline
    i = peak_idx_ccw
    while i > 0 and rads_ccw[i] < baseline_ccw:
        i -= 1
    
    if i == peak_idx_ccw:
        # never rose above baseline → use the peak angle
        baseline_dist_ccw = angs_ccw[peak_idx_ccw]
    else:
        # interpolate between points i and i+1
        r0, a0 = rads_ccw[i],   angs_ccw[i]
        r1, a1 = rads_ccw[i+1], angs_ccw[i+1]
        if not np.isfinite(r1 - r0) or r1 == r0:
            baseline_dist_ccw = 0.5 * (a0 + a1)
        else:
            frac = (baseline_ccw - r0) / (r1 - r0)
            baseline_dist_ccw = a0 + frac * (a1 - a0)
    
    peak_deg_ccw = angs_ccw[peak_idx_ccw]
    
    
    # 2) CW (clockwise) – search forward from the dent peak
    rads_cw, angs_cw = circum_cw.values, circum_cw.index.values
    dent_depth_cw   = nominal - rads_cw.min()
    baseline_cw_default     = nominal - baseline_circum * dent_depth_cw
    rads_cw_idx = find_outbound_deflection_end(rads_cw, nominal, slope_tolerance=0.004, circumferential_mode=True)
    # baseline_cw = rads_cw[rads_cw_idx] if rads_cw_idx is not None else baseline_cw_default
    baseline_cw = baseline_us
    # Re-establish the dent_depth_cw value based on the new baseline_cw value
    dent_depth_cw = baseline_cw - rads_cw.min()

    peak_idx_cw     = np.argmin(rads_cw)
    # walk forward until we reach or exceed the baseline
    i = peak_idx_cw
    while i < len(rads_cw) - 1 and rads_cw[i] < baseline_cw:
        i += 1
    
    if i == peak_idx_cw:
        baseline_dist_cw = angs_cw[peak_idx_cw]
    else:
        # interpolate between points i-1 and i
        r0, a0 = rads_cw[i-1], angs_cw[i-1]
        r1, a1 = rads_cw[i],   angs_cw[i]
        if not np.isfinite(r1 - r0) or r1 == r0:
            baseline_dist_cw = 0.5 * (a0 + a1)
        else:
            frac = (baseline_cw - r0) / (r1 - r0)
            baseline_dist_cw = a0 + frac * (a1 - a0)
    
    peak_deg_cw = angs_cw[peak_idx_cw]

    # LAX table header - counterclockwise
    sht['E27'] = 'Upstream Counterclockwise Lengths'
    sht['E28'] = 'Max Dent Depth'; sht['G28'] = dent_depth_ccw
    sht['E29'] = 'Baseline';       sht['F29'] = baseline_dist_ccw; sht['G29'] = baseline_ccw

    # LAX table header - clockwise
    sht['K27'] = 'Upstream Clockwise Lengths'
    sht['K28'] = 'Max Dent Depth'; sht['M28'] = dent_depth_cw
    sht['K29'] = 'Baseline';       sht['L29'] = baseline_dist_cw; sht['M29'] = baseline_cw
    
    # Prepare LTR% & endpoints - ccwstream
    percentages_ccw = [90,85,80,75,70,60,50,40,30,20,15,10]
    radii_ccw       = circum_ccw.values
    distances_ccw   = circum_ccw.index.values

    horiz_start = 45
    base_row    = horiz_start - 1  # for baseline endpoints
    
    # Prepare LTR% & endpoints - cwstream
    percentages_cw = [90,85,80,75,70,60,50,40,30,20,15,10]
    radii_cw       = circum_cw.values
    distances_cw   = circum_cw.index.values

    horiz_start = 45
    base_row    = horiz_start - 1  # for baseline endpoints

    # baseline points - ccwstream
    sht[f'E{base_row}'] = peak_deg_ccw
    sht[f'F{base_row}'] = baseline_dist_ccw
    sht[f'G{base_row}'] = baseline_ccw
    sht[f'H{base_row}'] = baseline_ccw

    targets_ccw = []
    for i, p in enumerate(percentages_ccw):
        # target_r_ccw = baseline_ccw - (p/100) * dent_depth_ccw
        target_r_ccw = rads_ccw.min() + (1 - p/100) * dent_depth_ccw
    
        # walk backward from the dent‐peak until radius >= target_r_ccw
        j = ce_pos
        while j > 0 and radii_ccw[j] < target_r_ccw:
            j -= 1
    
        if radii_ccw[j] < target_r_ccw:
            # never rose back to the target → use the peak location
            x_at = peak_deg_ccw
        else:
            # interpolate between j and j+1
            r0, a0 = radii_ccw[j],   distances_ccw[j]
            r1, a1 = radii_ccw[j+1], distances_ccw[j+1]
            if not np.isfinite(r1 - r0) or r1 == r0:
                x_at = 0.5*(a0 + a1)
            else:
                frac = (target_r_ccw - r0) / (r1 - r0)
                x_at = a0 + frac*(a1 - a0)
    
        LTR = abs(x_at - peak_deg_ccw)
        row = 30 + i
        sht[f'E{row}'] = f'LTR{p}%'
        sht[f'F{row}'] = LTR
        sht[f'G{row}'] = target_r_ccw

        # chart endpoints in columns H–K
        r = horiz_start + i
        sht[f'E{r}'] = peak_deg_ccw
        sht[f'F{r}'] = x_at
        sht[f'G{r}'] = target_r_ccw
        sht[f'H{r}'] = target_r_ccw

        targets_ccw.append((p, x_at))   
        
    # baseline endpoints - cwstream
    sht[f'K{base_row}'] = peak_deg_cw
    sht[f'L{base_row}'] = baseline_dist_cw
    sht[f'M{base_row}'] = baseline_cw
    sht[f'N{base_row}'] = baseline_cw

    targets_cw = []
    for i, p in enumerate(percentages_cw):
        # target_r_cw = baseline_cw - (p/100) * dent_depth_cw
        target_r_cw = rads_cw.min() + (1 - p/100) * dent_depth_cw
        # exact match?
        exact = np.isclose(radii_cw, target_r_cw)
        if exact.any():
            x_at = distances_cw[exact][0]
        else:
            # find first radius ≥ target
            above = np.where(radii_cw >= target_r_cw)[0]
            if above.size:
                j1 = above[0]
                j2 = max(j1-1, 0)
                x1, y1 = distances_cw[j2], radii_cw[j2]
                x2, y2 = distances_cw[j1], radii_cw[j1]
                if y2 == y1 or not np.isfinite(y2 - y1):
                    x_at = 0.5 * (x1 + x2)
                else:
                    x_at = x1 + (target_r_cw - y1) * (x2 - x1) / (y2 - y1)
            else:
                # never reaches target, use last point
                x_at = distances_cw[-1]
    
        LTR = abs(x_at - peak_deg_cw)
        row = 30 + i
        sht[f'K{row}'] = f'LTR{p}%'
        sht[f'L{row}'] = LTR
        sht[f'M{row}'] = target_r_cw

        r = horiz_start + i
        sht[f'K{r}'] = peak_deg_cw
        sht[f'L{r}'] = x_at
        sht[f'M{r}'] = target_r_cw
        sht[f'N{r}'] = target_r_cw
    
        targets_cw.append((p, x_at))

    # ccwstream Areas
    sht['E68'] = 'Counterclockwise Areas'
    sht['E69'] = 'Distance'
    sht['F69'] = 'Radius'
    start_ccw = min(baseline_dist_ccw, peak_deg_ccw)
    end_ccw   = max(baseline_dist_ccw, peak_deg_ccw)
    mask_seg_ccw = (circum_ccw.index >= start_ccw) & (circum_ccw.index <= end_ccw)
    seg_dist_ccw = circum_ccw.index[mask_seg_ccw].tolist()
    seg_rad_ccw  = circum_ccw.values[mask_seg_ccw].tolist()
    for i, value in enumerate(seg_dist_ccw):
        sht[f'E{70 + i}'] = value
    for i, value in enumerate(seg_rad_ccw):
        sht[f'F{70 + i}'] = value

    # cwstream Areas
    sht['K68'] = 'Clockwise Areas'
    sht['K69'] = 'Distance'
    sht['L69'] = 'Radius'
    start_cw = min(baseline_dist_cw, peak_deg_cw)
    end_cw   = max(baseline_dist_cw, peak_deg_cw)
    mask_seg_cw = (circum_cw.index >= start_cw) & (circum_cw.index <= end_cw)
    seg_dist_cw = circum_cw.index[mask_seg_cw].tolist()
    seg_rad_cw  = circum_cw.values[mask_seg_cw].tolist()
    for i, value in enumerate(seg_dist_cw):
        sht[f'K{70 + i}'] = value
    for i, value in enumerate(seg_rad_cw):
        sht[f'L{70 + i}'] = value

    # Area calc & store in list - counterclockwise
    sht['G69'] = 'Area'
    area_list_ccw = [0]
    for j in range(1, len(seg_dist_ccw)):
        d_i, d_im1 = seg_dist_ccw[j], seg_dist_ccw[j-1]
        r_i, r_im1 = seg_rad_ccw[j], seg_rad_ccw[j-1]
        area_ccw = 0.5 * (d_i - d_im1) * abs((r_i - baseline_ccw) + (r_im1 - baseline_ccw))
        area_list_ccw.append(area_ccw)
        sht[f'G{71 + j}'] = area_ccw

    # Area calc & store in list - clockwise
    sht['M69'] = 'Area'
    area_list_cw = [0]
    for j in range(1, len(seg_dist_cw)):
        d_i, d_im1 = seg_dist_cw[j], seg_dist_cw[j-1]
        r_i, r_im1 = seg_rad_cw[j], seg_rad_cw[j-1]
        area_cw = 0.5 * (d_i - d_im1) * abs((r_i - baseline_cw) + (r_im1 - baseline_cw))
        area_list_cw.append(area_cw)
        sht[f'M{71 + j}'] = area_cw

    # TR% cumulative areas - counterclockwise
    cum_row = 30
    for p in [85, 75, 60, 50, 40, 30, 20, 15, 10]:
        x_at_p = next(x for pct, x in targets_ccw if pct == p)
        idx_p  = min(range(len(seg_dist_ccw)), key=lambda k: abs(seg_dist_ccw[k] - x_at_p))
        cum_area = sum(area_list_ccw[idx_p:])
        sht[f'H{cum_row}'] = f'ATR{p}%'
        sht[f'I{cum_row}'] = cum_area
        cum_row += 1

    # TR% cumulative areas - clockwise
    cum_row = 30
    for p in [85,75,60,50,40,30,20,15,10]:
        x_at_p = next(x for pct, x in targets_cw if pct == p)
        idx_p  = min(range(len(seg_dist_cw)), key=lambda k: abs(seg_dist_cw[k] - x_at_p))
        # sum from baseline inward to the p% point
        cum_area = sum(area_list_cw[:idx_p+1])
        sht[f'N{cum_row}'] = f'ATR{p}%'
        sht[f'O{cum_row}'] = cum_area
        cum_row += 1

    # DS-CCW and DS-CW
    # Circumferential data
    # Create a new sheet titled "Circumferential Profiles", and assing to sht
    circumDS_name = 'CircumDS Profiles'
    sht = wb.create_sheet(title=circumDS_name)

    # Write the circum_df to sht without using writer to preserve Excel VBA macros
    for r_idx, row in enumerate(dataframe_to_rows(circum_df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            sht.cell(row=r_idx, column=c_idx, value=value)
    
    # --- 6.1 Compute interpolated baseline for CCW and CW ---
    
    # Common nominal
    nominal = nominal_IR
    
    # 1) CCW (counterclockwise) – search backward from the dent peak
    rads_ccw, angs_ccw = circum_ccw.values, circum_ccw.index.values
    dent_depth_ccw   = nominal - rads_ccw.min()
    baseline_ccw_default     = nominal - baseline_circum * dent_depth_ccw
    rads_ccw_idx = find_inbound_deflection_start(rads_ccw, nominal, slope_tolerance=0.004, circumferential_mode=True)
    # baseline_ccw = rads_ccw[rads_ccw_idx] if rads_ccw_idx is not None else baseline_ccw_default
    baseline_ccw = baseline_ds
    # Re-establish the dent_depth_ccw value based on the new baseline_ccw value
    dent_depth_ccw = baseline_ccw - rads_ccw.min()

    # find peak index
    peak_idx_ccw     = np.argmin(rads_ccw)
    # walk backward until we reach or exceed the baseline
    i = peak_idx_ccw
    while i > 0 and rads_ccw[i] < baseline_ccw:
        i -= 1
    
    if i == peak_idx_ccw:
        # never rose above baseline → use the peak angle
        baseline_dist_ccw = angs_ccw[peak_idx_ccw]
    else:
        # interpolate between points i and i+1
        r0, a0 = rads_ccw[i],   angs_ccw[i]
        r1, a1 = rads_ccw[i+1], angs_ccw[i+1]
        if not np.isfinite(r1 - r0) or r1 == r0:
            baseline_dist_ccw = 0.5 * (a0 + a1)
        else:
            frac = (baseline_ccw - r0) / (r1 - r0)
            baseline_dist_ccw = a0 + frac * (a1 - a0)
    
    peak_deg_ccw = angs_ccw[peak_idx_ccw]
    
    
    # 2) CW (clockwise) – search forward from the dent peak
    rads_cw, angs_cw = circum_cw.values, circum_cw.index.values
    dent_depth_cw   = nominal - rads_cw.min()
    baseline_cw_default     = nominal - baseline_circum * dent_depth_cw
    rads_cw_idx = find_outbound_deflection_end(rads_cw, nominal, slope_tolerance=0.004, circumferential_mode=True)
    # baseline_cw = rads_cw[rads_cw_idx] if rads_cw_idx is not None else baseline_cw_default
    baseline_cw = baseline_ds
    # Re-establish the dent_depth_cw value based on the new baseline_cw value
    dent_depth_cw = baseline_cw - rads_cw.min()

    peak_idx_cw     = np.argmin(rads_cw)
    # walk forward until we reach or exceed the baseline
    i = peak_idx_cw
    while i < len(rads_cw) - 1 and rads_cw[i] < baseline_cw:
        i += 1
    
    if i == peak_idx_cw:
        baseline_dist_cw = angs_cw[peak_idx_cw]
    else:
        # interpolate between points i-1 and i
        r0, a0 = rads_cw[i-1], angs_cw[i-1]
        r1, a1 = rads_cw[i],   angs_cw[i]
        if not np.isfinite(r1 - r0) or r1 == r0:
            baseline_dist_cw = 0.5 * (a0 + a1)
        else:
            frac = (baseline_cw - r0) / (r1 - r0)
            baseline_dist_cw = a0 + frac * (a1 - a0)
    
    peak_deg_cw = angs_cw[peak_idx_cw]

    # LAX table header - counterclockwise
    sht['E27'] = 'Downstream Counterclockwise Lengths'
    sht['E28'] = 'Max Dent Depth'; sht['G28'] = dent_depth_ccw
    sht['E29'] = 'Baseline';       sht['F29'] = baseline_dist_ccw; sht['G29'] = baseline_ccw

    # LAX table header - clockwise
    sht['K27'] = 'Downstream Clockwise Lengths'
    sht['K28'] = 'Max Dent Depth'; sht['M28'] = dent_depth_cw
    sht['K29'] = 'Baseline';       sht['L29'] = baseline_dist_cw; sht['M29'] = baseline_cw
    
    # Prepare LTR% & endpoints - ccwstream
    percentages_ccw = [90,85,80,75,70,60,50,40,30,20,15,10]
    radii_ccw       = circum_ccw.values
    distances_ccw   = circum_ccw.index.values

    horiz_start = 45
    base_row    = horiz_start - 1  # for baseline endpoints
    
    # Prepare LTR% & endpoints - cwstream
    percentages_cw = [90,85,80,75,70,60,50,40,30,20,15,10]
    radii_cw       = circum_cw.values
    distances_cw   = circum_cw.index.values

    horiz_start = 45
    base_row    = horiz_start - 1  # for baseline endpoints

    # baseline points - ccwstream
    sht[f'E{base_row}'] = peak_deg_ccw
    sht[f'F{base_row}'] = baseline_dist_ccw
    sht[f'G{base_row}'] = baseline_ccw
    sht[f'H{base_row}'] = baseline_ccw

    targets_ccw = []
    for i, p in enumerate(percentages_ccw):
        # target_r_ccw = baseline_ccw - (p/100) * dent_depth_ccw
        target_r_ccw = rads_ccw.min() + (1 - p/100) * dent_depth_ccw
    
        # walk backward from the dent‐peak until radius >= target_r_ccw
        j = ce_pos
        while j > 0 and radii_ccw[j] < target_r_ccw:
            j -= 1
    
        if radii_ccw[j] < target_r_ccw:
            # never rose back to the target → use the peak location
            x_at = peak_deg_ccw
        else:
            # interpolate between j and j+1
            r0, a0 = radii_ccw[j],   distances_ccw[j]
            r1, a1 = radii_ccw[j+1], distances_ccw[j+1]
            if not np.isfinite(r1 - r0) or r1 == r0:
                x_at = 0.5*(a0 + a1)
            else:
                frac = (target_r_ccw - r0) / (r1 - r0)
                x_at = a0 + frac*(a1 - a0)
    
        LTR = abs(x_at - peak_deg_ccw)
        row = 30 + i
        sht[f'E{row}'] = f'LTR{p}%'
        sht[f'F{row}'] = LTR
        sht[f'G{row}'] = target_r_ccw

        # chart endpoints in columns H–K
        r = horiz_start + i
        sht[f'E{r}'] = peak_deg_ccw
        sht[f'F{r}'] = x_at
        sht[f'G{r}'] = target_r_ccw
        sht[f'H{r}'] = target_r_ccw

        targets_ccw.append((p, x_at))   
        
    # baseline endpoints - cwstream
    sht[f'K{base_row}'] = peak_deg_cw
    sht[f'L{base_row}'] = baseline_dist_cw
    sht[f'M{base_row}'] = baseline_cw
    sht[f'N{base_row}'] = baseline_cw

    targets_cw = []
    for i, p in enumerate(percentages_cw):
        # target_r_cw = baseline_cw - (p/100) * dent_depth_cw
        target_r_cw = rads_cw.min() + (1 - p/100) * dent_depth_cw
        # exact match?
        exact = np.isclose(radii_cw, target_r_cw)
        if exact.any():
            x_at = distances_cw[exact][0]
        else:
            # find first radius ≥ target
            above = np.where(radii_cw >= target_r_cw)[0]
            if above.size:
                j1 = above[0]
                j2 = max(j1-1, 0)
                x1, y1 = distances_cw[j2], radii_cw[j2]
                x2, y2 = distances_cw[j1], radii_cw[j1]
                if y2 == y1 or not np.isfinite(y2 - y1):
                    x_at = 0.5 * (x1 + x2)
                else:
                    x_at = x1 + (target_r_cw - y1) * (x2 - x1) / (y2 - y1)
            else:
                # never reaches target, use last point
                x_at = distances_cw[-1]
    
        LTR = abs(x_at - peak_deg_cw)
        row = 30 + i
        sht[f'K{row}'] = f'LTR{p}%'
        sht[f'L{row}'] = LTR
        sht[f'M{row}'] = target_r_cw

        r = horiz_start + i
        sht[f'K{r}'] = peak_deg_cw
        sht[f'L{r}'] = x_at
        sht[f'M{r}'] = target_r_cw
        sht[f'N{r}'] = target_r_cw
    
        targets_cw.append((p, x_at))

    # ccwstream Areas
    sht['E68'] = 'Counterclockwise Areas'
    sht['E69'] = 'Distance'
    sht['F69'] = 'Radius'
    start_ccw = min(baseline_dist_ccw, peak_deg_ccw)
    end_ccw   = max(baseline_dist_ccw, peak_deg_ccw)
    mask_seg_ccw = (circum_ccw.index >= start_ccw) & (circum_ccw.index <= end_ccw)
    seg_dist_ccw = circum_ccw.index[mask_seg_ccw].tolist()
    seg_rad_ccw  = circum_ccw.values[mask_seg_ccw].tolist()
    for i, value in enumerate(seg_dist_ccw):
        sht[f'E{70 + i}'] = value
    for i, value in enumerate(seg_rad_ccw):
        sht[f'F{70 + i}'] = value

    # cwstream Areas
    sht['K68'] = 'Clockwise Areas'
    sht['K69'] = 'Distance'
    sht['L69'] = 'Radius'
    start_cw = min(baseline_dist_cw, peak_deg_cw)
    end_cw   = max(baseline_dist_cw, peak_deg_cw)
    mask_seg_cw = (circum_cw.index >= start_cw) & (circum_cw.index <= end_cw)
    seg_dist_cw = circum_cw.index[mask_seg_cw].tolist()
    seg_rad_cw  = circum_cw.values[mask_seg_cw].tolist()
    for i, value in enumerate(seg_dist_cw):
        sht[f'K{70 + i}'] = value
    for i, value in enumerate(seg_rad_cw):
        sht[f'L{70 + i}'] = value

    # Area calc & store in list - counterclockwise
    sht['G69'] = 'Area'
    area_list_ccw = [0]
    for j in range(1, len(seg_dist_ccw)):
        d_i, d_im1 = seg_dist_ccw[j], seg_dist_ccw[j-1]
        r_i, r_im1 = seg_rad_ccw[j], seg_rad_ccw[j-1]
        area_ccw = 0.5 * (d_i - d_im1) * abs((r_i - baseline_ccw) + (r_im1 - baseline_ccw))
        area_list_ccw.append(area_ccw)
        sht[f'G{71 + j}'] = area_ccw

    # Area calc & store in list - clockwise
    sht['M69'] = 'Area'
    area_list_cw = [0]
    for j in range(1, len(seg_dist_cw)):
        d_i, d_im1 = seg_dist_cw[j], seg_dist_cw[j-1]
        r_i, r_im1 = seg_rad_cw[j], seg_rad_cw[j-1]
        area_cw = 0.5 * (d_i - d_im1) * abs((r_i - baseline_cw) + (r_im1 - baseline_cw))
        area_list_cw.append(area_cw)
        sht[f'M{71 + j}'] = area_cw

    # TR% cumulative areas - counterclockwise
    cum_row = 30
    for p in [85, 75, 60, 50, 40, 30, 20, 15, 10]:
        x_at_p = next(x for pct, x in targets_ccw if pct == p)
        idx_p  = min(range(len(seg_dist_ccw)), key=lambda k: abs(seg_dist_ccw[k] - x_at_p))
        cum_area = sum(area_list_ccw[idx_p:])
        sht[f'H{cum_row}'] = f'ATR{p}%'
        sht[f'I{cum_row}'] = cum_area
        cum_row += 1

    # TR% cumulative areas - clockwise
    cum_row = 30
    for p in [85,75,60,50,40,30,20,15,10]:
        x_at_p = next(x for pct, x in targets_cw if pct == p)
        idx_p  = min(range(len(seg_dist_cw)), key=lambda k: abs(seg_dist_cw[k] - x_at_p))
        # sum from baseline inward to the p% point
        cum_area = sum(area_list_cw[:idx_p+1])
        sht[f'N{cum_row}'] = f'ATR{p}%'
        sht[f'O{cum_row}'] = cum_area
        cum_row += 1

    # --- 8) Shape Results: 4‑panel summary (directly reference sheet cells) ---
    sht_res = wb.create_sheet('Shape Results')
    
    # Titles
    start_circUS = 2 + len(percentages_us)
    start_circDS = 2 + start_circUS + len(percentages_ccw)
    sht_res['A1'] = 'Axial - Upstream'
    sht_res['F1'] = 'Axial - Downstream'
    sht_res[f'A{start_circUS}'] = 'Upstream Circumferential - Counterclockwise'
    sht_res[f'F{start_circUS}'] = 'Upstream Circumferential - Clockwise'
    sht_res[f'A{start_circDS}'] = 'Downstream Circumferential - Counterclockwise'
    sht_res[f'F{start_circDS}'] = 'Downstream Circumferential - Clockwise'

    # Upstream/Downstream
    # copy the twelve LAX% labels
    for i, p in enumerate(percentages_us):
        row = 2 + i
        sht_res[f'A{row}'] = f'LAX{p}%'
        sht_res[f'F{row}'] = f'LAX{p}%'  # same percentages for downstream

    # copy the nine AX% labels
    for i, p in enumerate([85,75,60,50,40,30,20,15,10]):
        row = 2 + i
        sht_res[f'C{row}'] = f'AX{p}%'
        sht_res[f'H{row}'] = f'AX{p}%'
    
    # now formulas to pull the values
    for i in range(len(percentages_us)):
        src_row = 30 + i
        dst_row = 2 + i
    
        # Upstream LAX from Axial!F30:F41 → ShapeResults!B2:B13
        sht_res[f'B{dst_row}'] = f"='{axial_name}'!F{src_row}"

        # Downstream LAX from Axial!L30:L41 → ShapeResults!G2:G13
        sht_res[f'G{dst_row}'] = f"='{axial_name}'!L{src_row}"
    
    for i in range(9):
        src_row = 30 + i
        dst_row = 2 + i
    
        # Upstream AX from Axial!I30:I38 → ShapeResults!D2:D10
        sht_res[f'D{dst_row}'] = f"='{axial_name}'!I{src_row}"

        # Downstream AX from Axial!O30:O38 → ShapeResults!H2:H10
        sht_res[f'I{dst_row}'] = f"='{axial_name}'!O{src_row}"

    # US-CCW/CW
    # copy the twelve LAX% labels
    for i, p in enumerate(percentages_ccw):
        row = start_circUS + 1 + i
        sht_res[f'A{row}'] = f'LTR{p}%'
        sht_res[f'F{row}'] = f'LTR{p}%'

    # copy the nine AX% labels
    for i, p in enumerate([85,75,60,50,40,30,20,15,10]):
        row = start_circUS + 1 + i
        sht_res[f'C{row}'] = f'ATR{p}%'
        sht_res[f'H{row}'] = f'ATR{p}%'

    for i, p in enumerate(percentages_ccw):
        src_row = 30 + i
        dst_row = start_circUS + 1 + i
    
        # pull CCW length from Circumferential!F30:F41 → ShapeResults!B{start_circUS+…}
        sht_res[f'B{dst_row}'] = f"='{circumUS_name}'!F{src_row}"

        # pull CW  length from Circumferential!L30:L41 → ShapeResults!G{start_circUS+…}
        sht_res[f'G{dst_row}'] = f"='{circumUS_name}'!L{src_row}"

    # areas for CCW (cells I30:I38) and CW (cells O30:O38)
    for i in range(9):
        src_row = 30 + i
        dst_row_cc = start_circUS + 1 + i

        # CCW area → ShapeResults!C{dst_row_cc}
        sht_res[f'D{dst_row_cc}'] = f"='{circumUS_name}'!I{src_row}"

        # CW area → ShapeResults!H{dst_row_cc}
        sht_res[f'I{dst_row_cc}'] = f"='{circumUS_name}'!O{src_row}"

    # DS-CCW/CW
    # copy the twelve LAX% labels
    for i, p in enumerate(percentages_ccw):
        row = start_circDS + 1 + i
        sht_res[f'A{row}'] = f'LTR{p}%'
        sht_res[f'F{row}'] = f'LTR{p}%'

    # copy the nine AX% labels
    for i, p in enumerate([85,75,60,50,40,30,20,15,10]):
        row = start_circDS + 1 + i
        sht_res[f'C{row}'] = f'ATR{p}%'
        sht_res[f'H{row}'] = f'ATR{p}%'

    for i, p in enumerate(percentages_ccw):
        src_row = 30 + i
        dst_row = start_circDS + 1 + i
    
        # pull CCW length from Circumferential!F30:F41 → ShapeResults!B{start_circDS+…}
        sht_res[f'B{dst_row}'] = f"='{circumDS_name}'!F{src_row}"

        # pull CW  length from Circumferential!L30:L41 → ShapeResults!G{start_circDS+…}
        sht_res[f'G{dst_row}'] = f"='{circumDS_name}'!L{src_row}"

    # areas for CCW (cells I30:I38) and CW (cells O30:O38)
    for i in range(9):
        src_row = 30 + i
        dst_row_cc = start_circDS + 1 + i

        # CCW area → ShapeResults!C{dst_row_cc}
        sht_res[f'D{dst_row_cc}'] = f"='{circumDS_name}'!I{src_row}"

        # CW area → ShapeResults!H{dst_row_cc}
        sht_res[f'I{dst_row_cc}'] = f"='{circumDS_name}'!O{src_row}"

    #######################################################################################
    # Add graphs
    #######################################################################################
    # Axial plot using a Solid Line for the Upstream data, and a Dashed line for the Downstream data
    sht = wb[axial_name]
    chart = ScatterChart()
    chart.title = "Dent Axial Profile"
    chart.x_axis.title = "Axial Distance (in)"
    chart.y_axis.title = "Radius (in)"
    chart.width = 24
    chart.height = 10

    # Upstream series
    values = Reference(sht, min_col=2, min_row=3, max_row=2+len(axial_df))
    xvalues = Reference(sht, min_col=1, min_row=3, max_row=2+len(axial_df))
    series1 = Series(values, xvalues, title_from_data=False)
    series1.title = SeriesLabel(v="Upstream")
    chart.series.append(series1)

    # Downstream series
    values2 = Reference(sht, min_col=3, min_row=3, max_row=2+len(axial_df))
    series2 = Series(values2, xvalues, title_from_data=False)
    series2.title = SeriesLabel(v="Downstream")
    series2.graphicalProperties.ln = LineProperties(prstDash='dash')
    chart.series.append(series2)

    # Configure chart legend and axes
    chart.legend.position = 'r'  # or use LegendPos.r if available
    chart.legend.overlay = False
    chart.x_axis.visible = True
    chart.y_axis.visible = True
    chart.x_axis.majorTickMark = 'out'
    chart.y_axis.majorTickMark = 'out'

    # Insert chart at E2
    sht.add_chart(chart, "E2")

    # Zoomed in view axial plot the same as the previous, plus additional lines for the Baseline and LAX lines
    chart = ScatterChart()
    chart.title = "Axial Lengths"
    chart.x_axis.title = "Axial Distance (in)"
    chart.y_axis.title = "Radius (in)"
    chart.width = 24
    chart.height = 10
    # Set the x axis to min of baseline_dist_ds and max of baseline_dist_us
    chart.x_axis.min = baseline_dist_us
    chart.x_axis.max = baseline_dist_ds

    # Add series for Upstream and Downstream
    chart.series.append(series1)
    chart.series.append(series2)

    # Add series for the Baseline-US and Baseline-DS
    values_BUS = Reference(sht, min_col=7, min_row=base_row, max_col=8)
    xvalues_BUS = Reference(sht, min_col=5, min_row=base_row, max_col=6)
    series_BUS = Series(values_BUS, xvalues_BUS, title_from_data=False)
    series_BUS.title = SeriesLabel(v="Baseline-US")
    chart.series.append(series_BUS)

    values_BDS = Reference(sht, min_col=13, min_row=base_row, max_col=14)
    xvalues_BDS = Reference(sht, min_col=11, min_row=base_row, max_col=12)
    series_BDS = Series(values_BDS, xvalues_BDS, title_from_data=False)
    series_BDS.title = SeriesLabel(v="Baseline-DS")
    series_BDS.graphicalProperties.line.dash_style = 'dot'
    chart.series.append(series_BDS)

    # Add series for all of the LAX% lines, with the color being defined based on the palette selection
    for p, _ in targets_us:
        color = palette[p % len(palette)]
        r = horiz_start + percentages_us.index(p)
        val = Reference(sht, min_col=7, min_row=r, max_col=8)
        xval = Reference(sht, min_col=5, min_row=r, max_col=6)
        val_series = Series(val, xval, title_from_data=False)
        val_series.title = SeriesLabel(v=f"LAX{p}%-US")
        # val_series.graphicalPropertiesline.solidFill = Color(rgb=color)
        chart.series.append(val_series)

    for p, _ in targets_ds:
        color = palette[p % len(palette)]
        r = horiz_start + percentages_ds.index(p)
        val = Reference(sht, min_col=13, min_row=r, max_col=14)
        xval = Reference(sht, min_col=11, min_row=r, max_col=12)
        val_series = Series(val, xval, title_from_data=False)
        val_series.title = SeriesLabel(v=f"LAX{p}%-DS")
        # val_series.graphicalPropertiesline.solidFill = Color(rgb=color)
        val_series.graphicalProperties.ln = LineProperties(prstDash='dash')
        chart.series.append(val_series)

    # Configure chart legend and axes
    chart.legend.position = 'r'  # or use LegendPos.r if available
    chart.legend.overlay = False
    chart.x_axis.visible = True
    chart.y_axis.visible = True
    chart.x_axis.majorTickMark = 'out'
    chart.y_axis.majorTickMark = 'out'

    # Insert chart at E43
    sht.add_chart(chart, "E43")

    # Upstream - Circumferential plot using a Solid Line for the Counterclockwise data, and a Dashed line for the Clockwise data
    sht = wb[circumUS_name]
    chart = ScatterChart()
    chart.title = "Dent Circumferential Profile"
    chart.x_axis.title = "Circumferential Distance (in)"
    chart.y_axis.title = "Radius (in)"
    chart.width = 24
    chart.height = 10

    # Upstream series
    values = Reference(sht, min_col=2, min_row=3, max_row=2+len(axial_df))
    xvalues = Reference(sht, min_col=1, min_row=3, max_row=2+len(axial_df))
    series1 = Series(values, xvalues, title_from_data=False)
    series1.title = SeriesLabel(v="Counterclockwise")
    chart.series.append(series1)

    # Downstream series
    values2 = Reference(sht, min_col=3, min_row=3, max_row=2+len(axial_df))
    series2 = Series(values2, xvalues, title_from_data=False)
    series2.title = SeriesLabel(v="Clockwise")
    series2.graphicalProperties.ln = LineProperties(prstDash='dash')
    chart.series.append(series2)

    # Configure chart legend and axes
    chart.legend.position = 'r'  # or use LegendPos.r if available
    chart.legend.overlay = False
    chart.x_axis.visible = True
    chart.y_axis.visible = True
    chart.x_axis.majorTickMark = 'out'
    chart.y_axis.majorTickMark = 'out'

    # Insert chart at E2
    sht.add_chart(chart, "E2")

    # Zoomed in view circumferential plot the same as the previous, plus additional lines for the Baseline and LAX lines
    chart = ScatterChart()
    chart.title = "Upstream Circumferential Lengths"
    chart.x_axis.title = "Circumferential Distance (in)"
    chart.y_axis.title = "Radius (in)"
    chart.width = 24
    chart.height = 10
    # Set the x axis to min of baseline_dist_cw and max of baseline_dist_ccw
    chart.x_axis.min = baseline_dist_cw
    chart.x_axis.max = baseline_dist_ccw

    # Add series for CCW and CW
    chart.series.append(series1)
    chart.series.append(series2)

    # Add series for the Baseline-CCW and Baseline-CW
    values_BCCW = Reference(sht, min_col=7, min_row=base_row, max_col=8)
    xvalues_BCCW = Reference(sht, min_col=5, min_row=base_row, max_col=6)
    series_BCCW = Series(values_BCCW, xvalues_BCCW, title_from_data=False)
    series_BCCW.title = SeriesLabel(v="Baseline-USCCW")
    chart.series.append(series_BCCW)

    values_BCW = Reference(sht, min_col=13, min_row=base_row, max_col=14)
    xvalues_BCW = Reference(sht, min_col=11, min_row=base_row, max_col=12)
    series_BCW = Series(values_BCW, xvalues_BCW, title_from_data=False)
    series_BCW.title = SeriesLabel(v="Baseline-USCW")
    series_BCW.graphicalProperties.line.dash_style = 'dot'
    chart.series.append(series_BCW)

    # Add series for all of the LAX% lines, with the color being defined based on the palette selection
    for p, _ in targets_ccw:
        color = palette[p % len(palette)]
        r = horiz_start + percentages_ccw.index(p)
        val = Reference(sht, min_col=7, min_row=r, max_col=8)
        xval = Reference(sht, min_col=5, min_row=r, max_col=6)
        val_series = Series(val, xval, title_from_data=False)
        val_series.title = SeriesLabel(v=f"LAX{p}%-CCW")
        # val_series.graphicalPropertiesline.solidFill = Color(rgb=color)
        chart.series.append(val_series)

    for p, _ in targets_cw:
        color = palette[p % len(palette)]
        r = horiz_start + percentages_cw.index(p)
        val = Reference(sht, min_col=13, min_row=r, max_col=14)
        xval = Reference(sht, min_col=11, min_row=r, max_col=12)
        val_series = Series(val, xval, title_from_data=False)
        val_series.title = SeriesLabel(v=f"LAX{p}%-CW")
        # val_series.graphicalPropertiesline.solidFill = Color(rgb=color)
        val_series.graphicalProperties.ln = LineProperties(prstDash='dash')
        chart.series.append(val_series)

    # Configure chart legend and axes
    chart.legend.position = 'r'  # or use LegendPos.r if available
    chart.legend.overlay = False
    chart.x_axis.visible = True
    chart.y_axis.visible = True
    chart.x_axis.majorTickMark = 'out'
    chart.y_axis.majorTickMark = 'out'

    # Insert chart at E43
    sht.add_chart(chart, "E43")

    # Downstrea - Circumferential plot using a Solid Line for the Counterclockwise data, and a Dashed line for the Clockwise data
    sht = wb[circumDS_name]
    chart = ScatterChart()
    chart.title = "Dent Circumferential Profile"
    chart.x_axis.title = "Circumferential Distance (in)"
    chart.y_axis.title = "Radius (in)"
    chart.width = 24
    chart.height = 10

    # Upstream series
    values = Reference(sht, min_col=2, min_row=3, max_row=2+len(axial_df))
    xvalues = Reference(sht, min_col=1, min_row=3, max_row=2+len(axial_df))
    series1 = Series(values, xvalues, title_from_data=False)
    series1.title = SeriesLabel(v="Counterclockwise")
    chart.series.append(series1)

    # Downstream series
    values2 = Reference(sht, min_col=3, min_row=3, max_row=2+len(axial_df))
    series2 = Series(values2, xvalues, title_from_data=False)
    series2.title = SeriesLabel(v="Clockwise")
    series2.graphicalProperties.ln = LineProperties(prstDash='dash')
    chart.series.append(series2)

    # Configure chart legend and axes
    chart.legend.position = 'r'  # or use LegendPos.r if available
    chart.legend.overlay = False
    chart.x_axis.visible = True
    chart.y_axis.visible = True
    chart.x_axis.majorTickMark = 'out'
    chart.y_axis.majorTickMark = 'out'

    # Insert chart at E2
    sht.add_chart(chart, "E2")

    # Zoomed in view circumferential plot the same as the previous, plus additional lines for the Baseline and LAX lines
    chart = ScatterChart()
    chart.title = "Downstream Circumferential Lengths"
    chart.x_axis.title = "Circumferential Distance (in)"
    chart.y_axis.title = "Radius (in)"
    chart.width = 24
    chart.height = 10
    # Set the x axis to min of baseline_dist_cw and max of baseline_dist_ccw
    chart.x_axis.min = baseline_dist_cw
    chart.x_axis.max = baseline_dist_ccw

    # Add series for CCW and CW
    chart.series.append(series1)
    chart.series.append(series2)

    # Add series for the Baseline-CCW and Baseline-CW
    values_BCCW = Reference(sht, min_col=7, min_row=base_row, max_col=8)
    xvalues_BCCW = Reference(sht, min_col=5, min_row=base_row, max_col=6)
    series_BCCW = Series(values_BCCW, xvalues_BCCW, title_from_data=False)
    series_BCCW.title = SeriesLabel(v="Baseline-DSCCW")
    chart.series.append(series_BCCW)

    values_BCW = Reference(sht, min_col=13, min_row=base_row, max_col=14)
    xvalues_BCW = Reference(sht, min_col=11, min_row=base_row, max_col=12)
    series_BCW = Series(values_BCW, xvalues_BCW, title_from_data=False)
    series_BCW.title = SeriesLabel(v="Baseline-DSCW")
    series_BCW.graphicalProperties.line.dash_style = 'dot'
    chart.series.append(series_BCW)

    # Add series for all of the LAX% lines, with the color being defined based on the palette selection
    for p, _ in targets_ccw:
        color = palette[p % len(palette)]
        r = horiz_start + percentages_ccw.index(p)
        val = Reference(sht, min_col=7, min_row=r, max_col=8)
        xval = Reference(sht, min_col=5, min_row=r, max_col=6)
        val_series = Series(val, xval, title_from_data=False)
        val_series.title = SeriesLabel(v=f"LAX{p}%-CCW")
        # val_series.graphicalPropertiesline.solidFill = Color(rgb=color)
        chart.series.append(val_series)

    for p, _ in targets_cw:
        color = palette[p % len(palette)]
        r = horiz_start + percentages_cw.index(p)
        val = Reference(sht, min_col=13, min_row=r, max_col=14)
        xval = Reference(sht, min_col=11, min_row=r, max_col=12)
        val_series = Series(val, xval, title_from_data=False)
        val_series.title = SeriesLabel(v=f"LAX{p}%-CW")
        # val_series.graphicalPropertiesline.solidFill = Color(rgb=color)
        val_series.graphicalProperties.ln = LineProperties(prstDash='dash')
        chart.series.append(val_series)

    # Configure chart legend and axes
    chart.legend.position = 'r'  # or use LegendPos.r if available
    chart.legend.overlay = False
    chart.x_axis.visible = True
    chart.y_axis.visible = True
    chart.x_axis.majorTickMark = 'out'
    chart.y_axis.majorTickMark = 'out'

    # Insert chart at E43
    sht.add_chart(chart, "E43")

    # Save and close the Excel workbook
    wb.save(file_path)
    wb.close()