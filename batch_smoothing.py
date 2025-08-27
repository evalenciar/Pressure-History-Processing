"""
Python script that uses processing module that performs the following:
- Load the Pipe Tally containing all feature metadata
- Iterate through the .csv/.xlsx documents in a folder, searching for matching feature names
- Load the feature data and perform data smoothing using processing.smooth_data
- Export the smoothed data (and metadata) to the existing results Excel document for each feature saved in another directory
"""

import pandas as pd
import numpy as np
import time
import os
from tkinter import Tk, filedialog, simpledialog
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from matplotlib import pyplot as plt

import processing as pro

# Default file paths to eliminate prompts
pipe_tally = r"C:\Users\emman\OneDrive - Softnostics\Projects\100001 - 100025\100004 (Acuren - Southbow Screening Software)\Client Documents\KS12-14 Data Collection - (Fixed Headers).xlsx"

# Function that iterates through a list of .csv/.xlsx files, checks for matching feature names, and processes the data
def iterate_files(file_list, output_list, category, pipe_tally_df):
    
    for file_path in file_list:
        start_time = time.time()
        # Separate the filename from the file_path
        filename = os.path.basename(file_path)
        # Typical pattern matching for feature ID involves a number after the string "Dent" in the filename
        # However, this number may have several trailing zeros and be followed by a period or underscore. Separate just the number.
        try:
            feature_id = int(filename.split("Dent")[-1].split(".")[0].split("_")[0].strip(" ").lstrip("0"))
        except ValueError:
            print(f"{category} Feature {feature_id}: Could not extract feature ID from filename: '{filename}'")
            with open(os.path.join(output_folder, f"{category}_RLA_Results.txt"), "a") as f:
                f.write(f"{category} Feature {feature_id}: Could not extract feature ID from filename: '{filename}'\n")
            continue
        # Filter pipe_tally_df Section header for string having category, then search for Feature ID header matching feature_id
        try:
            df_feature = pipe_tally_df[(pipe_tally_df['Section'].astype(str).str.contains(category, case=False, na=False)) & (pipe_tally_df['Feature ID'] == feature_id)]
        except Exception as e:
            print(f"{category} Feature {feature_id}: Error finding feature in the Pipe Tally: {e}")
            with open(os.path.join(output_folder, f"{category}_RLA_Results.txt"), "a") as f:
                f.write(f"{category} Feature {feature_id}: Error finding feature in the Pipe Tally: {e}\n")
            continue
        # Extract OD, WT, SMYS from the df_feature. Convert units to imperial
        try:
            OD = df_feature['TCPL NPS'].values[0] # Units in inches
            WT = df_feature['TCPL Nominal Wall Thickness [mm]'].values[0] * 0.0393701  # Convert from mm to inches
            SMYS = df_feature['TCPL SMYS [MPa]'].values[0] * 145.038  # Convert from MPa to psi
        except Exception as e:
            print(f"{category} Feature {feature_id}: Error extracting data for feature in the Pipe Tally: {e}")
            with open(os.path.join(output_folder, f"{category}_RLA_Results.txt"), "a") as f:
                f.write(f"{category} Feature {feature_id}: Error extracting data for feature in the Pipe Tally: {e}\n")
            continue
        # If category is KS12, use ILI_format = 'KS12', otherwise use ILI_format = 'KS1314'
        if category == 'KS12':
            ILI_format = 'KS12'
        else:
            ILI_format = 'KS1314'
        # Create a Process class for each file
        df = pro.Process(file_path, ILI_format, OD, WT, SMYS, filename)
        # Smooth the data
        df.smooth_data()
        # Create a new DataFrame containing smoothed data
        df_smoothed = pd.DataFrame(data=df.f_radius, index=df.f_axial, columns=df.f_circ)

        # Find one folder directory in the output_list that contains matching substring of feature_id
        # The folders in the output_list will always have format "Feature X"
        matching_folders = [f for f in output_list if str(feature_id) == os.path.basename(f).split("Feature ")[-1]]
        if not matching_folders:
            print(f"{category} Feature {feature_id}: No matching output folder found for feature")
            with open(os.path.join(output_folder, f"{category}_RLA_Results.txt"), "a") as f:
                f.write(f"{category} Feature {feature_id}: No matching output folder found for feature\n")
            continue
        output_folder = matching_folders[0]
        print(f"{category} Feature {feature_id}: Using output folder: '{os.path.basename(output_folder)}'")

        # Create a 2x1 figure having contour line plots of: raw and smooth data
        # Establish the levels as the absolute maximum and minimum of both the raw and smooth radius values
        # levels = np.linspace(min([df.o_radius.min(), df.f_radius.min()]), max([df.o_radius.max(), df.f_radius.max()]), 17)
        levels_raw = np.linspace(df.o_radius.min(), df.o_radius.max(), 17)
        levels_smooth = np.linspace(df.f_radius.min(), df.f_radius.max(), 17)

        fig, axs = plt.subplots(2, 1, figsize=(10, 8), sharex=True)
        # Raw data contour plot
        c1 = axs[0].contourf(df.o_axial, df.o_circ, df.o_radius.T, levels=levels_raw, cmap='viridis')
        cb1 = fig.colorbar(c1, ax=axs[0])
        cb1.set_label("Radius (inches)")
        axs[0].set_ylabel(f"{category} Feature {feature_id}: Raw Data")
        axs[0].set_yticks([])

        # Smoothed data contour plot
        c2 = axs[1].contourf(df.f_axial, df.f_circ, df.f_radius.T, levels=levels_smooth, cmap='viridis')
        cb2 = fig.colorbar(c2, ax=axs[1])
        cb2.set_label("Radius (inches)")
        axs[1].set_xlabel("Axial Position (inches)")
        axs[1].set_ylabel(f"{category} Feature {feature_id}: Smoothed Data")
        axs[1].set_yticks([])

        plt.suptitle(f"Comparison for {category} Feature {feature_id}", fontsize=16)
        plt.tight_layout()
        plt.savefig(os.path.join(output_folder, f"{category}_Feature_{feature_id}_Contour.jpg"), dpi=300)
        plt.close(fig)

        # In the output_folder, find an .xlsm Excel document containing substring "Feature RLA", keep absolute path for the xlsm files
        xlsm_files = [os.path.join(output_folder, f) for f in os.listdir(output_folder) if f.endswith('.xlsm') and 'Feature RLA' in f]
        if not xlsm_files:
            print(f"{category} Feature {feature_id}: No matching .xlsm file found in output folder: '{os.path.basename(output_folder)}'")
            with open(os.path.join(output_folder, f"{category}_RLA_Results.txt"), "a") as f:
                f.write(f"{category} Feature {feature_id}: No matching .xlsm file found in output folder: '{os.path.basename(output_folder)}'\n")
            continue
        xlsm_file = xlsm_files[0]
        # Save the df_smoothed as a new sheet in the existing .xlsm file
        wb = openpyxl.load_workbook(xlsm_file, read_only=False, keep_vba=True)
        # Remove the sheet if it already exists (optional, for clean overwrite)
        if "Smoothed Data" in wb.sheetnames:
            wb.remove(wb["Smoothed Data"])
        # Create a new sheet named "Smoothed Data"
        ws = wb.create_sheet(title="Smoothed Data")
        for r in dataframe_to_rows(df_smoothed, index=True, header=True):
            ws.append(r)
        # Create a new sheet named "Pipe Tally" and save the df_feature data
        ws_tally = wb.create_sheet(title="Pipe Tally")
        for r in dataframe_to_rows(df_feature, index=False, header=True):
            ws_tally.append(r)
        # Save as a new filename "<category>_Feature_<feature_id>_Results.xlsm"
        wb.save(os.path.join(output_folder, f"{category}_Feature_{feature_id}_Results.xlsm"))

        print(f"{category} Feature {feature_id}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully saved smoothed data to file: '{os.path.basename(xlsm_file)}'")

        # Save results to notepad after each iteration. Keep the notepad in the parent output folder (one directory higher)
        with open(os.path.join(os.path.dirname(output_folder), f"{category}_Smoothing_Results.txt"), "a") as f:
            f.write(f"{category} Feature {feature_id}: ({round(time.time() - start_time)}s/{round(time.time() - overall_start_time)}s) Successfully saved smoothed data to file: '{os.path.basename(xlsm_file)}'\n")

def main():
    # Hide the main Tkinter window
    root = Tk()
    root.withdraw()

    # Keep track of overall time
    global overall_start_time
    overall_start_time = time.time()

    # Split the workflow into three categories
    categories = ['KS12', 'KS13', 'KS14']

    # Prompt the user to select the Pipe Tally Excel file, and load the sheet "Sheet1" into a DataFrame
    pipe_tally_path = filedialog.askopenfilename(title="Select Pipe Tally Excel file", filetypes=[("Excel files", "*.xlsx;*.xls")])
    if not pipe_tally_path:
        print("No file selected. Exiting script.")
        exit()
    else:
        print(f"Selected file: {pipe_tally_path}")

    try:
        pipe_tally_df = pd.read_excel(pipe_tally_path, sheet_name="Sheet1", header=1, engine='calamine')
        print("Pipe Tally loaded successfully.")
    except Exception as e:
        print(f"Error loading Pipe Tally: {e}")
        exit()

    # Filter the pipe_tally_df to include only rows where 'Feature Type' is Dent
    pipe_tally_df = pipe_tally_df[pipe_tally_df['Feature Type'] == 'Dent']

    ####################################################################
    # Select Results Output Excel Documents
    ####################################################################

    # Prompt the user to select the output folder containing the output result Excel documents
    output_folder = filedialog.askdirectory(title="Select Output Folder")
    if not output_folder:
        print("No folder selected. Exiting script.")
        exit()
    else:
        print(f"Selected folder: {output_folder}")

    # Check if the output folder contains any sub folders that contain a substring matching with categories
    output_subfolders = [f.path for f in os.scandir(output_folder) if f.is_dir() and any(cat in f.name for cat in categories)]
    if not output_subfolders:
        print("No subfolders found for the specified categories. Exiting script.")
        exit()
    else:
        print(f"Found subfolders for outputs.")

    ####################################################################
    # Select Raw Data
    ####################################################################

    # Prompt the user to select the raw data folder containing the .csv/.xlsx files
    raw_folder = filedialog.askdirectory(title="Select Raw Data Folder")
    if not raw_folder:
        print("No folder selected. Exiting script.")
        exit()
    else:
        print(f"Selected folder: {raw_folder}")

    # Check if the raw data folder contains any sub folders for the categories
    raw_subfolders = [f.path for f in os.scandir(raw_folder) if f.is_dir() and any(cat in f.name for cat in categories)]
    if not raw_subfolders:
        # No raw_subfolders found, check for files directly in the raw_folder
        print("No subfolders found for the specified categories, proceed with files in the main folder.")
        csv_files = [f.path for f in os.scandir(raw_folder) if f.is_file() and f.name.endswith(('.csv', '.xlsx'))]
        if not csv_files:
            print("No .csv or .xlsx files found. Exiting script.")
            exit()
        else:
            print(f"Found .csv/.xlsx files.")
            # 
    else:
        print(f"Found subfolders for raw data.")
        # Iterate through each category
        for category in categories:
            # Select the corresponding output_subfolder if it contains a substring matching with the category
            output_subfolder = next((f for f in output_subfolders if category in f), None)
            if output_subfolder is None:
                print(f"No output subfolder found for category: {category}. Skipping.")
                continue
            # Select the corresponding raw_subfolder if it contains a substring matching with the category
            raw_subfolder = next((f for f in raw_subfolders if category in f), None)
            if raw_subfolder is None:
                print(f"No raw subfolder found for category: {category}. Skipping.")
                continue
            # Begin processing the data for the current category. Only show the folder base name
            print(f"Begin processing for category: '{category}', using output folder: '{os.path.basename(output_subfolder)}', and raw folder: '{os.path.basename(raw_subfolder)}'")
            # Using the selected folders, iterate through the features in the folder, make sure to keep absolute path
            iterate_files([os.path.join(raw_subfolder, f) for f in os.listdir(raw_subfolder)], [os.path.join(output_subfolder, f) for f in os.listdir(output_subfolder)], category, pipe_tally_df)

# Run script when loaded
if __name__ == "__main__":
    main()
