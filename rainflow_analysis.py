"""
Rainflow Analysis
The objective of this script is to perform a rainflow analysis on pressure history data.
"""

# ============================================================================
# ONLY CHANGE THESE VARIABLES FOR DIFFERENT VERSIONS OF RLA TEMPLATE
rla_template_folder = 'templates'
rla_template_name = 'RLA (v1.8.3).xlsm'
# ============================================================================

import numpy as np
import pandas as pd
import rainflow
import matplotlib.pyplot as plt
import openpyxl
import warnings
import math
import os

rla_template = os.path.join(rla_template_folder, rla_template_name)

# Ignore all UserWarnings from openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Turn off interactive plotting
plt.ioff()  

# =============================================================================
# Functions
# =============================================================================

class DentData:
    def __init__(self, dent_category, dent_ID, OD, WT, SMYS, MAOP, service_years, M, min_range, Lx, hx, SG, L1, L2, h1, h2, D1, D2, confidence, CPS, interaction_weld, interaction_corrosion, dent_depth_percent, ili_pressure, restraint_condition, ml_depth_percent, ml_location, orientation, vendor_comments):
        self.dent_category = dent_category
        self.dent_ID = dent_ID
        self.OD = OD
        self.WT = WT
        self.SMYS = SMYS
        self.MAOP = MAOP
        self.service_years = service_years
        self.M = M
        self.min_range = min_range
        self.Lx = Lx
        self.hx = hx
        self.SG = SG
        self.L1 = L1
        self.L2 = L2
        self.h1 = h1
        self.h2 = h2
        self.D1 = D1
        self.D2 = D2
        self.confidence = confidence
        self.CPS = CPS
        self.interaction_weld = interaction_weld
        self.interaction_corrosion = interaction_corrosion
        self.dent_depth_percent = dent_depth_percent
        self.ili_pressure = ili_pressure
        self.restraint_condition = restraint_condition
        self.ml_depth_percent = ml_depth_percent
        self.ml_location = ml_location
        self.orientation = orientation
        self.vendor_comments = vendor_comments

def liquid(P_list, P_time, results_path, dd, press_dict: dict = None, save_history:bool=False, save_cycles:bool=False, save_md49:bool=False, create_excel:bool=True):
    # Extract the information for the specified dent
    dent_category = dd.dent_category
    dent_ID = dd.dent_ID
    OD = dd.OD
    WT = dd.WT
    SMYS = dd.SMYS
    MAOP = dd.MAOP
    service_years = dd.service_years
    M = dd.M
    min_range = dd.min_range
    Lx = dd.Lx
    hx = dd.hx
    SG = dd.SG
    L1 = dd.L1
    L2 = dd.L2
    h1 = dd.h1
    h2 = dd.h2
    D1 = dd.D1
    D2 = dd.D2
    
    # Determine the operational pressures at the dent location. Taken from Equation (5) in API 1183 Section 6.6.3.1 Rainflow Counting
    P = Px(Lx, hx, P_list[0], P_list[1], SG, L1, L2, h1, h2, D1, D2)

    # Determine the mean pressure as %SMYS where s = PD / (2t) [%]
    P_mean_smys = float(100 * (np.average(P) * OD / (2 * WT)) / SMYS)

    # Rainflow Analysis using Python package 'rainflow'. Output is in format: Pressure Range [psig], Pressure Mean [psig], Cycle Count, Index Start, Index End
    cycles = pd.DataFrame(rainflow.extract_cycles(P)).to_numpy()

    # Filter the cycles array based on the min_range using the Pressure Range (first column)
    cycles = cycles[cycles[:, 0] > min_range]

    # Calculate the SSI, CI, and SSI MD49
    SSI = equivalent_cycles('ssi', cycles, OD, WT, SMYS, service_years, M, min_range)
    CI = equivalent_cycles('ci', cycles, OD, WT, SMYS, service_years, M, min_range)
    MD49_SSI, MD49_bins = MD49(cycles, OD, WT, SMYS, service_years, M, min_range)
    if press_dict:
        Neq_SSI, bin_cycles = custom_bins(cycles, OD, WT, SMYS, service_years, M, press_dict, min_range)
    else:
        Neq_SSI, bin_cycles = None, None

    if save_history:
        df_P = pd.DataFrame(data=P, columns=['Pressure (psig)'], index=P_time)
        df_P.to_csv(results_path + f"Feature {dent_ID} " + 'Interpolated_Pressure_History_Data.csv', header=False, index=False)

    if save_cycles:
        # Save the cycles to a .txt file
        np.savetxt(results_path + f"Feature {dent_ID} " + 'cycles.csv', cycles, delimiter=',')

    if save_md49:
        # Save the MD49_bins to a .txt file
        np.savetxt(results_path + f"Feature {dent_ID} " + 'md49_bins.csv', MD49_bins, delimiter=',')

    if create_excel:
        create_RLA_Excel(results_path, dd, cycles, MD49_bins, P_mean_smys)
    
    # Graphing
    liquid_graphing(dent_ID, results_path, P, P_time, dent_category)
    
    return SSI, CI, MD49_SSI, cycles, MD49_bins, Neq_SSI, bin_cycles
  
def liquid_graphing(dent_ID, results_path, P, P_time, dent_category:str = ''):
    # Save the interpolated pressure history 
    fig, sp = plt.subplots(figsize=(8,4), dpi=240)
    fig.suptitle(f'Pressure History for {dent_category} Feature {dent_ID}', fontsize=16)
    sp.scatter(P_time, P, s=0.1)
    sp.grid(color='lightgray', alpha=0.5, zorder=1)
    sp.set_ylim([0, max(1750, math.ceil(np.nanmax(P) / 250) * 250)]) # Set y-axis limit to a maximum of 2000 or the next multiple of 250 above the max pressure
    sp.set_ylabel('Interpolated Pressure (psig)')
    sp.set_xlabel('Date Time')
    fig.savefig(os.path.join(results_path, f"{dent_category}_Feature_{dent_ID}_Interpolated_Pressure_History.png"))
    plt.close(fig)

def create_RLA_Excel(results_path, dd, cycles, MD49_bins, P_mean_smys):
    dent_category = dd.dent_category
    dent_ID = dd.dent_ID
    OD = dd.OD 
    WT = dd.WT 
    SMYS = dd.SMYS
    MAOP = dd.MAOP
    service_years = dd.service_years
    min_range = dd.min_range
    confidence = dd.confidence
    CPS = dd.CPS
    interaction_weld = dd.interaction_weld
    interaction_corrosion = dd.interaction_corrosion
    dent_depth_percent = dd.dent_depth_percent
    ili_pressure = dd.ili_pressure
    restraint_condition = dd.restraint_condition
    ml_depth_percent = dd.ml_depth_percent
    ml_location = dd.ml_location

    # Defaults:
    start_row = 3
    rainflow_column = 1 # Column A
    md49_column = 37    # Column AK
    
    ref_path = rla_template
    wb = openpyxl.load_workbook(filename=ref_path, read_only=False, keep_vba=True)
    
    # Update the values in the Summary tab
    wbs = wb['Summary']
    wbs['D4'] = round(OD, 3)
    wbs['D5'] = round(WT, 3)
    wbs['D6'] = round(SMYS, -3) # Round to nearest 1000 so we get 70,000
    wbs['D8'] = round(MAOP, 0)
    wbs['D9'] = round(service_years, 3)
    wbs['D10'] = float(min_range)

    wbs['H4'] = str(dent_category)
    wbs['H5'] = str(dent_ID)

    wbs['B72'] = rla_template_name # Use this to keep track of file version
    
    # Import the values in the Rainflow tab (begins on A3)
    wbs = wb['Rainflow']
    wbs['A1'] = f'Rainflow Analysis Outputs with Filter > {round(min_range,1)} psig'
    for row_i, row_val in enumerate(cycles):
        for col_i, _ in enumerate(row_val):
            wbs.cell(row=start_row + row_i, column=rainflow_column + col_i).value = float(cycles[row_i, col_i])

    # Import the MD49_bins to the MD49 section (begins on AK3)
    wbs = wb['Rainflow']
    for row_i, row_val in enumerate(MD49_bins):
        wbs.cell(row=start_row + row_i, column=md49_column).value = float(MD49_bins[row_i])

    # Save the P_mean_smys
    wbs['K61'] = float(ili_pressure)
    wbs['K63'] = float(dent_depth_percent)
    wbs['K65'] = float(P_mean_smys)
    wbs['K67'] = str(interaction_corrosion)
    wbs['K68'] = float(ml_depth_percent)
    wbs['K69'] = str(ml_location)
    wbs['K70'] = str(interaction_weld)
    wbs['K72'] = float(confidence)
    wbs['K73'] = str(CPS)
    wbs['K84'] = str(restraint_condition)

    # Save the resultant Excel workbook into the designated folder
    wb.save(filename=os.path.join(results_path, f"{dent_category}_Feature_{dent_ID}_Results.xlsm"))
    wb.close()

def Px(Lx,hx,P1,P2,SG,L1,L2,h1,h2,D1,D2):
    """
    Taken from API 1176
    
    Note: the version in API 1183 Section 6.6.3.1 Rainflow Counting Equation (5) is incorrect.
    
    Parameters
    ----------
    Lx : float
        the location of point analysis, ft
    hx : float
        the elevation of point analysis, ft
    P1 : array
        the upstream discharge pressure, psig
    P2 : array
        the downstream suction pressure, psig
    K : float
        SG x (0.433 psi/ft), where SG = specific gravity of product
    L1 : float
        the location of upstream discharge station, ft
    L2 : float
        the location of downstream suction station, ft
    h1 : float
        the elevation of upstream discharge station, ft
    h2 : float
        the elevation of downstream suction station, ft
    D1 : float
        the pipe diameter of segment between L1 and Lx, in
    D2 : float
        the pipe diameter of segment between Lx and L2, in

    Returns
    -------
    The intermediate pressure point between pressure sources, psig

    """
    K = SG * 0.433  # Convert SG to psi/ft
    Px = (P1 + K*h1 - P2 - K*h2)*(1/(((Lx - L1)*D2**5)/((L2 - Lx)*D1**5) + 1)) - K*(hx - h2) + P2
    
    return Px

def custom_bins(cycles, OD: float, WT: float, SMYS: float, service_years: float, M: float, press_dict: dict, min_range: float =5) -> tuple[float, dict]:
    '''
    Use the custom pressure bins defined in press_dict to sum the cycles into the bins.
    Parameters
    ----------
    cycles : array of floats
        the array output from the rainflow analysis, with columns: [Pressure Range (psig), Pressure Mean (psig), Cycle Count, Index Start, Index End]
    OD : float
        the outside diameter of the pipe, in
    WT : float
        the wall thickness of the pipe, in
    SMYS : float
        the specified minimum yield strength of the pipe, psi
    service_years : float
        the number of years the pressure history represents, years
    M : float
        the slope of the S-N curve, typically 3.0 for steel
    press_dict : dict
        dictionary containing pressure bin values (pmin, pmax, prange, pmean) in %SMYS
    min_range : float, optional
        the minimum pressure range to consider for the analysis, default is 5 psi

    Returns
    -------
    (bin_SSI : float, bin_cycles : np.ndarray)
        A tuple containing the total damage equivalent cycles and a numpy array of the custom bins with their respective cycle counts.
    '''
    # Filter out cycles below the minimum range
    custom_cycles = cycles.copy()
    for i in range(len(custom_cycles)):
        if custom_cycles[i,0] < min_range: 
            custom_cycles[i,2] = 0
    # Convert pressure values into units of % SMYS
    custom_cycles[:,0] = 100*custom_cycles[:,0]*OD/(2*WT)/SMYS
    custom_cycles[:,1] = 100*custom_cycles[:,1]*OD/(2*WT)/SMYS
    
    # Make the second level keys lowercase for consistency, and the first level keys integers
    press_dict = {int(k): {kk.lower(): vv for kk, vv in v.items()} for k, v in press_dict.items()}
    # Create an empty dictionary to hold the cycle counts for each bin
    bin_cycles = {k: 0 for k in press_dict.keys()}
    
    # Create groups of equivalent Prange bins which will then be used for the iterative processing
    bin_groups = {}
    for bin_num, bin_vals in press_dict.items():
        if bin_vals['prange'] not in bin_groups:
            bin_groups[bin_vals['prange']] = []
        bin_groups[bin_vals['prange']].append({int(bin_num): bin_vals["pmean"]})
    # Sort the bin_groups by the prange key
    bin_groups = dict(sorted(bin_groups.items()))

    # Iterate through every pressure range cycle, and find the appropriate bin to add the cycle count to using the bin_groups
    for i, press_range in enumerate(custom_cycles[:, 0]):
        for bin_range, bin_vals in bin_groups.items():
            if press_range <= bin_range:
                for i, bin_val in enumerate(bin_vals):
                    if i != len(bin_vals) - 1 and (custom_cycles[i,1] <= (list(bin_vals[i].values())[0] + list(bin_vals[i + 1].values())[0])/2):  # If the mean is less than the midpoint between this bin and the next bin, add it here
                        bin_cycles[list(bin_val.keys())[0]] += custom_cycles[i,2]
                        break
                    if i == len(bin_vals) - 1:  # Last bin, so just add it here
                        bin_cycles[list(bin_val.keys())[0]] += custom_cycles[i,2]
                        break
                break

    # # Inversed version of bin_groups, where the keys are the bin means and the values are dicts of bin numbers and their pranges
    # bin_groups2 = {}
    # for bin_num, bin_vals in press_dict.items():
    #     if bin_vals['pmean'] not in bin_groups2:
    #         bin_groups2[bin_vals['pmean']] = []
    #     bin_groups2[bin_vals['pmean']].append({int(bin_num): bin_vals["prange"]})
    # # Sort the bin_groups2 by the bin means (the keys), and also sort the inner lists by their pranges
    # for key in bin_groups2:
    #     bin_groups2[key] = sorted(bin_groups2[key], key=lambda x: list(x.values())[0])
    # bin_groups = dict(sorted(bin_groups2.items()))

    # # Iterate through every pressure range cycle, and find the appropriate bin to add the cycle count to using the bin_groups
    # for i, press_mean in enumerate(custom_cycles[:, 1]):
    #     for bin_mean, bin_vals in bin_groups.items():
    #         if press_mean <= bin_mean:
    #             for i, bin_val in enumerate(bin_vals):
    #                 if i != len(bin_vals) - 1 and (custom_cycles[i,1] <= (list(bin_vals[i].values())[0] + list(bin_vals[i + 1].values())[0])/2):  # If the mean is less than the midpoint between this bin and the next bin, add it here
    #                     bin_cycles[list(bin_val.keys())[0]] += custom_cycles[i,2]
    #                     break
    #                 if i == len(bin_vals) - 1:  # Last bin, so just add it here
    #                     bin_cycles[list(bin_val.keys())[0]] += custom_cycles[i,2]
    #                     break
    #             break

    # Calculate the damage equivalent cycles for the custom bins
    SSI_ref_stress = 13000  # psi
    # The Neq SSI = (Prange_%SMYS * SMYS / SSI_ref_stress) ^ M * Cycles
    Prange_pct_smys = np.array([v['prange'] for v in press_dict.values()])
    Neq_SSI = sum(((Prange_pct_smys* SMYS / SSI_ref_stress) ** M) * np.array(list(bin_cycles.values()))) / service_years
    # FOR REFERENCE: Calculate the MD49 Equivalent Cycles
    # Neq_SSI_MD49s = sum(((((MD49_P_range/100)*SMYS)/SSI_ref_stress)**M)*MD49_bins)/service_years
    # return Neq_SSI, np.array(list(bin_cycles.values()))
    return Neq_SSI, bin_cycles

def MD49(cycles, OD, WT, SMYS, service_years, M, min_range=5, identifier=''):
    # Create an empty array for all the MD-4-9 bins
    MD49_bins = np.zeros(28)
    MD49_P_range = np.array([10,20,30,40,50,60,70,
                             10,20,30,40,50,60,
                             10,20,30,40,50,
                             10,20,30,40,
                             10,20,30,
                             10,20,
                             10])
    
    # Remove any MD49_cycles that have a pressure range below the minimum range
    MD49_cycles = cycles.copy()
    for i, val in enumerate(MD49_cycles[:,0]):
        if MD49_cycles[i,0] < min_range: 
            MD49_cycles[i,2] = 0
    
    # Reference Stress Ranges
    SSI_ref_stress = 13000  # psi
    SSI_ref_press = SSI_ref_stress*2*WT/OD
    
    # Convert pressure range into units of % SMYS
    MD49_cycles[:,0] = 100*MD49_cycles[:,0]*OD/(2*WT)/SMYS
    MD49_cycles[:,1] = 100*MD49_cycles[:,1]*OD/(2*WT)/SMYS
    
    # Iterate through every pressure range cycle
    for i, press_range in enumerate(MD49_cycles[:, 0]):
        # Pressure range: 0 - 10% SMYS
        if MD49_cycles[i, 0] <= 10.0: #if range is 0 - 10% SMYS
            if MD49_cycles[i, 1] <= 20.0: #if mean is 0 - 20% SMYS
                MD49_bins[0] = MD49_bins[0] + MD49_cycles[i, 2] #BIN #1
            elif MD49_cycles[i, 1] <= 30.0: #if mean is 20 - 30% SMYS
                MD49_bins[7] = MD49_bins[7] + MD49_cycles[i, 2] #BIN #8
            elif MD49_cycles[i, 1] <= 40.0: #if mean is 30 - 40% SMYS
                MD49_bins[13] = MD49_bins[13] + MD49_cycles[i, 2] #BIN #14
            elif MD49_cycles[i, 1] <= 50.0: #if mean is 40 - 50% SMYS
                MD49_bins[18] = MD49_bins[18] + MD49_cycles[i, 2] #BIN #19
            elif MD49_cycles[i, 1] <= 60.0: #if mean is 50 - 60% SMYS
                MD49_bins[22] = MD49_bins[22] + MD49_cycles[i, 2] #BIN #23
            elif MD49_cycles[i, 1] <= 70.0: #if mean is 60 - 70% SMYS
                MD49_bins[25] = MD49_bins[25] + MD49_cycles[i, 2] #BIN #26
            else: #if mean is >70% SMYS
                MD49_bins[27] = MD49_bins[27] + MD49_cycles[i, 2] #BIN #28
        # Pressure range: 10 - 20% SMYS
        elif MD49_cycles[i, 0] <= 20.0: #if range is 10 - 20% SMYS
            if MD49_cycles[i, 1] <= 25.0: #if mean is 0 - 25% SMYS
                MD49_bins[1] = MD49_bins[1] + MD49_cycles[i, 2] #BIN #2
            elif MD49_cycles[i, 1] <= 35.0: #if mean is 25 - 35% SMYS
                MD49_bins[8] = MD49_bins[8] + MD49_cycles[i, 2] #BIN #9
            elif MD49_cycles[i, 1] <= 45.0: #if mean is 35 - 45% SMYS
                MD49_bins[14] = MD49_bins[14] + MD49_cycles[i, 2] #BIN #15
            elif MD49_cycles[i, 1] <= 55.0: #if mean is 45 - 55% SMYS
                MD49_bins[19] = MD49_bins[19] + MD49_cycles[i, 2] #BIN #20
            elif MD49_cycles[i, 1] <= 65.0: #if mean is 55 - 65% SMYS
                MD49_bins[23] = MD49_bins[23] + MD49_cycles[i, 2] #BIN #24
            else: #if mean is >65% SMYS
                MD49_bins[26] = MD49_bins[26] + MD49_cycles[i, 2] #BIN #27
        # Pressure range: 20 - 30% SMYS
        elif MD49_cycles[i, 0] <= 30.0: #if range is 20 - 30% SMYS
            if MD49_cycles[i, 1] <= 30.0: #if mean is 0 - 30% SMYS
                MD49_bins[2] = MD49_bins[2] + MD49_cycles[i, 2] #BIN #3
            elif MD49_cycles[i, 1] <= 40.0: #if mean is 30 - 40% SMYS
                MD49_bins[9] = MD49_bins[9] + MD49_cycles[i, 2] #BIN #10
            elif MD49_cycles[i, 1] <= 50.0: #if mean is 40 - 50% SMYS
                MD49_bins[15] = MD49_bins[15] + MD49_cycles[i, 2] #BIN #16
            elif MD49_cycles[i, 1] <= 60.0: #if mean is 50 - 60% SMYS
                MD49_bins[20] = MD49_bins[20] + MD49_cycles[i, 2] #BIN #21
            else: #if mean is >60% SMYS
                MD49_bins[24] = MD49_bins[24] + MD49_cycles[i, 2] #BIN #25
        # Pressure range: 30 - 40% SMYS
        elif MD49_cycles[i, 0] <= 40.0: #if range is 30 - 40% SMYS
            if MD49_cycles[i, 1] <= 35.0: #if mean is 0 - 35% SMYS
                MD49_bins[3] = MD49_bins[3] + MD49_cycles[i, 2] #BIN #4
            elif MD49_cycles[i, 1] <= 45.0: #if mean is 35 - 45% SMYS
                MD49_bins[10] = MD49_bins[10] + MD49_cycles[i, 2] #BIN #11
            elif MD49_cycles[i, 1] <= 55.0: #if mean is 45 - 55% SMYS
                MD49_bins[16] = MD49_bins[16] + MD49_cycles[i, 2] #BIN #17
            else: #if mean is >55% SMYS
                MD49_bins[21] = MD49_bins[21] + MD49_cycles[i, 2] #BIN #22
        # Pressure range: 40 - 50% SMYS
        elif MD49_cycles[i, 0] <= 50.0: #if range is 40 - 50% SMYS
            if MD49_cycles[i, 1] <= 40.0: #if mean is 0 - 40% SMYS
                MD49_bins[4] = MD49_bins[4] + MD49_cycles[i, 2] #BIN #5
            elif MD49_cycles[i, 1] <= 50.0: #if mean is 40 - 50% SMYS
                MD49_bins[11] = MD49_bins[11] + MD49_cycles[i, 2] #BIN #12
            else: #if mean is >50% SMYS
                MD49_bins[17] = MD49_bins[17] + MD49_cycles[i, 2] #BIN #18
        # Pressure range: 50 - 60% SMYS
        elif MD49_cycles[i, 0] <= 60.0: #if range is 50 - 60% SMYS
            if MD49_cycles[i, 1] <= 45.0: #if mean is 0 - 45% SMYS
                MD49_bins[5] = MD49_bins[5] + MD49_cycles[i, 2] #BIN #6
            else: #if mean is >45% SMYS
                MD49_bins[12] = MD49_bins[12] + MD49_cycles[i, 2] #BIN #13
        # Pressure range > 60% SMYS
        else: #if pressure range > 60% SMYS
            MD49_bins[6] = MD49_bins[6] + MD49_cycles[i, 2] #BIN #7
            
    # Calculate the MD49 Equivalent Cycles
    MD49_final_cycles = sum(((((MD49_P_range/100)*SMYS)/SSI_ref_stress)**M)*MD49_bins)/service_years
    
    return MD49_final_cycles, MD49_bins

def equivalent_cycles(index, cycles, OD, WT, SMYS, service_years, M, min_range=5):
    """
    Parameters
    ----------
    index : string
        SSI or CI. SSI is the Spectrum Severity Indicator, CI is the Cyclic Index
    cycles : array of floats
        the array output from the rainflow analysis
    service_years : float
        the period of time in years for the pressure history data
    min_range : array
        the threshold value for pressure ranges to consider. Default is 5 psi.

    Returns
    -------
    Either the SSI or CI. 

    """
    equiv_cycles = np.zeros(cycles.shape[0])
    
    # Reference Stress Ranges
    SSI_ref_stress = 13000  # psi
    SSI_ref_press = SSI_ref_stress*2*WT/OD
    CI_ref_stress = 37580   # psi
    CI_ref_press = CI_ref_stress*2*WT/OD
    
    if index.lower() == 'ssi':
        ref_press = SSI_ref_press
    elif index.lower() == 'ci':
        ref_press = CI_ref_press
        
    for i, val in enumerate(cycles[:,0]):
        if cycles[i,0] > min_range: 
            equiv_cycles[i] = ((cycles[i,0]/ref_press)**M)*cycles[i,2]
        else:
            equiv_cycles[i] = 0
            
    num_cycles = sum(equiv_cycles)/service_years
    
    return num_cycles